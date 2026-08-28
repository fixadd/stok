from __future__ import annotations

from datetime import date, datetime, timedelta
from collections import Counter
from uuid import uuid4

import mimetypes
import os
import shutil
import tempfile
import subprocess

from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urlparse

from flask import (
    Flask,
    abort,
    after_this_request,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from sqlalchemy import func, text
from sqlalchemy.orm import joinedload
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import load_workbook
from .navigation import build_breadcrumbs, build_sidebar_sections
from .routes.admin import register_admin_routes
from .routes.auth import register_auth_routes
from .routes.inventory import register_inventory_routes
from .routes.maintenance import register_maintenance_routes
from .routes.information import register_information_routes
from .routes.profile import register_profile_routes
from .routes.requests import register_request_routes
from .routes.stock import register_stock_routes

from .services.authz import (
    current_actor_name,
    get_active_user,
    get_system_role,
    has_system_role,
    is_safe_redirect_target,
    set_active_user,
)
from .utils.parsing import (
    parse_excel_date,
    parse_int_or_none,
    sanitize_input_text,
    sanitize_metadata_payload,
)
from .personnel_lifecycle import personnel_lifecycle_bp
from .models import (
    Brand,
    Factory,
    HardwareModel,
    HardwareType,
    InfoCategory,
    InfoAttachment,
    InfoEntry,
    InventoryEvent,
    InventoryItem,
    InventoryLicense,
    InventoryMaintenance,
    LdapProfile,
    LicenseName,
    ProductCatalogEntry,
    RequestGroup,
    RequestLine,
    RequestLineSnapshot,
    RequestOrder,
    UsageArea,
    User,
    db,
    find_existing_by_name,
    ActivityLog,
    StockCategory,
    StockItem,
    StockLog,
    StockMovement,
    StockUnit,
    StockAssignment,
    StockAuditLog,
)

INVENTORY_STATUSES = {"aktif", "beklemede", "arizali", "hurda", "stokta"}
DEFAULT_EVENT_ACTOR = "Sistem"
LICENSE_STATUS_LABELS = {
    "aktif": "Aktif",
    "pasif": "Pasif",
    "beklemede": "Beklemede",
}


STOCK_CATEGORY_LABELS = {
    "envanter": "Envanter",
    "cevre_birimi": "Çevre Birimi",
    "yazici": "IP Yazıcı",
    "lisans": "Lisans",
    "talep": "Talep",
    "manuel": "Manuel",
}

STOCK_STATUS_LABELS = {
    "stokta": "Stokta",
    "devredildi": "Devredildi",
    "arizali": "Arızalı",
    "hurda": "Hurda",
}

STOCK_STATUS_CLASSES = {
    "stokta": "status-stock",
    "devredildi": "status-assigned",
    "arizali": "status-faulty",
    "hurda": "status-scrap",
}

STOCK_SOURCE_LABELS = {
    "inventory": "Envanter Takip",
    "license": "Lisans Takip",
    "request": "Talep Takip",
    "manual": "Manuel Kayıt",
}


SYSTEM_ROLE_LEVELS = {
    "user": 0,
    "admin": 1,
    "superadmin": 2,
}

SYSTEM_ROLE_LABELS = {
    "user": "Kullanıcı",
    "admin": "Admin",
    "superadmin": "Süper Admin",
}

INFO_ALLOWED_EXTENSIONS: dict[str, set[str]] = {
    ".png": {"image/png"},
    ".jpg": {"image/jpeg"},
    ".jpeg": {"image/jpeg"},
    ".gif": {"image/gif"},
    ".webp": {"image/webp"},
    ".bmp": {"image/bmp"},
    ".pdf": {"application/pdf"},
    ".txt": {"text/plain"},
    ".csv": {"text/csv", "application/csv", "application/vnd.ms-excel"},
    ".doc": {
        "application/msword",
        "application/vnd.ms-word",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    },
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    },
    ".xls": {
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    },
    ".ppt": {
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    },
    ".pptx": {
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/vnd.ms-powerpoint",
    },
}

INLINE_SAFE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
DISALLOWED_DOWNLOAD_EXTENSIONS = {".html", ".htm", ".svg", ".js", ".mjs"}
MAX_INFO_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB


STOCK_METADATA_FIELDS: dict[str, list[dict[str, Any]]] = {
    "envanter": [
        {
            "key": "inventory_no",
            "label": "Envanter No",
            "placeholder": "ENV-001",
            "required": True,
        },
        {
            "key": "hardware_type",
            "label": "Donanım Tipi",
            "placeholder": "Örn. Dizüstü Bilgisayar",
            "required": True,
        },
        {
            "key": "brand",
            "label": "Marka",
            "placeholder": "Marka",
            "required": True,
        },
        {
            "key": "model",
            "label": "Model",
            "placeholder": "Model",
            "required": True,
        },
        {
            "key": "serial_no",
            "label": "Seri No",
            "placeholder": "Seri numarası",
            "required": False,
        },
        {
            "key": "computer_name",
            "label": "Cihaz Adı",
            "placeholder": "Örn. IT-LAPTOP-01",
            "required": False,
        },
        {
            "key": "factory",
            "label": "Fabrika",
            "placeholder": "Fabrika adı",
            "required": True,
            "assignment_only": True,
            "options_key": "factories",
        },
        {
            "key": "department",
            "label": "Departman",
            "placeholder": "Departman",
            "required": True,
            "assignment_only": True,
            "options_key": "departments",
        },
        {
            "key": "responsible",
            "label": "Sorumlu",
            "placeholder": "Sorumlu kişi",
            "required": True,
            "assignment_only": True,
            "options_key": "responsibles",
        },
        {
            "key": "ifs_no",
            "label": "IFS No",
            "placeholder": "IFS-00001",
            "required": False,
            "assignment_only": True,
        },
    ],
    "cevre_birimi": [
        {
            "key": "hardware_type",
            "label": "Donanım Tipi",
            "placeholder": "Örn. Klavye",
            "required": True,
        },
        {
            "key": "brand",
            "label": "Marka",
            "placeholder": "Marka",
            "required": False,
        },
        {
            "key": "model",
            "label": "Model",
            "placeholder": "Model",
            "required": False,
        },
        {
            "key": "serial_no",
            "label": "Seri No",
            "placeholder": "Seri numarası",
            "required": False,
        },
        {
            "key": "factory",
            "label": "Fabrika",
            "placeholder": "Fabrika adı",
            "required": False,
            "assignment_only": True,
            "options_key": "factories",
        },
        {
            "key": "department",
            "label": "Departman",
            "placeholder": "Departman",
            "required": False,
            "assignment_only": True,
            "options_key": "departments",
        },
        {
            "key": "responsible",
            "label": "Sorumlu",
            "placeholder": "Teslim edilen kişi",
            "required": True,
            "assignment_only": True,
            "options_key": "responsibles",
        },
    ],
    "yazici": [
        {
            "key": "inventory_no",
            "label": "Envanter No",
            "placeholder": "IPY-001",
            "required": True,
        },
        {
            "key": "brand",
            "label": "Marka",
            "placeholder": "Marka",
            "required": True,
        },
        {
            "key": "model",
            "label": "Model",
            "placeholder": "Model",
            "required": True,
        },
        {
            "key": "serial_no",
            "label": "Seri No",
            "placeholder": "Seri numarası",
            "required": False,
        },
        {
            "key": "usage_area",
            "label": "Kullanım Alanı",
            "placeholder": "Örn. Finans",
            "required": False,
            "assignment_only": True,
            "options_key": "usage_areas",
        },
        {
            "key": "factory",
            "label": "Fabrika",
            "placeholder": "Fabrika adı",
            "required": True,
            "assignment_only": True,
            "options_key": "factories",
        },
        {
            "key": "hostname",
            "label": "Hostname",
            "placeholder": "PRN-OFIS-01",
            "required": False,
            "assignment_only": True,
        },
        {
            "key": "ip_address",
            "label": "IP Adresi",
            "placeholder": "10.0.0.10",
            "required": False,
            "assignment_only": True,
        },
        {
            "key": "mac_address",
            "label": "MAC Adresi",
            "placeholder": "AA:BB:CC:DD:EE:FF",
            "required": False,
            "assignment_only": True,
        },
        {
            "key": "responsible",
            "label": "Sorumlu",
            "placeholder": "Sorumlu kişi",
            "required": True,
            "assignment_only": True,
            "options_key": "responsibles",
        },
    ],
    "lisans": [
        {
            "key": "license_name",
            "label": "Lisans Adı",
            "placeholder": "Ürün adı",
            "required": True,
            "options_key": "license_names",
        },
        {
            "key": "license_key",
            "label": "Lisans Anahtarı",
            "placeholder": "XXXX-XXXX-XXXX",
            "required": True,
        },
        {
            "key": "inventory_no",
            "label": "Bağlı Envanter",
            "placeholder": "ENV-001",
            "required": False,
            "options_key": "inventory_numbers",
        },
        {
            "key": "factory",
            "label": "Fabrika",
            "placeholder": "Fabrika adı",
            "required": False,
            "assignment_only": True,
            "options_key": "factories",
        },
        {
            "key": "department",
            "label": "Departman",
            "placeholder": "Departman",
            "required": False,
            "assignment_only": True,
            "options_key": "departments",
        },
        {
            "key": "responsible",
            "label": "Sorumlu",
            "placeholder": "Teslim edilen kişi",
            "required": False,
            "assignment_only": True,
            "options_key": "responsibles",
        },
    ],
    "talep": [
        {
            "key": "hardware_type",
            "label": "Donanım Tipi",
            "placeholder": "Donanım tipi",
            "required": True,
        },
        {
            "key": "brand",
            "label": "Marka",
            "placeholder": "Marka",
            "required": False,
        },
        {
            "key": "model",
            "label": "Model",
            "placeholder": "Model",
            "required": False,
        },
        {
            "key": "department",
            "label": "Departman",
            "placeholder": "Departman",
            "required": False,
        },
    ],
    "manuel": [
        {
            "key": "hardware_type",
            "label": "Donanım Tipi",
            "placeholder": "Donanım tipi",
            "required": True,
        },
        {
            "key": "brand",
            "label": "Marka",
            "placeholder": "Marka",
            "required": False,
        },
        {
            "key": "model",
            "label": "Model",
            "placeholder": "Model",
            "required": False,
        },
    ],
}


THEME_OPTIONS = {
    "varsayilan": {
        "label": "Varsayılan",
        "description": "Hafif mavi tonlarda, modern varsayılan görünüm.",
        "preview": {"bg": "#eef4ff", "fg": "#1f2933"},
    },
    "gece": {
        "label": "Gece Modu",
        "description": "Koyu arka plan ve yüksek kontrastlı metinler.",
        "preview": {"bg": "#111827", "fg": "#f9fafb"},
    },
    "okyanus": {
        "label": "Okyanus",
        "description": "Serin mavi ve turkuaz geçişleriyle dinlendirici bir tema.",
        "preview": {"bg": "#0f172a", "fg": "#38bdf8"},
    },
    "orman": {
        "label": "Orman",
        "description": "Yeşil tonlarda doğal ve sakin bir görünüm.",
        "preview": {"bg": "#0b3d2e", "fg": "#c3f0ca"},
    },
    "gunes": {
        "label": "Güneş",
        "description": "Sıcak sarı ve turuncu vurgularla enerjik bir tema.",
        "preview": {"bg": "#fff7ed", "fg": "#c2410c"},
    },
    "lavanta": {
        "label": "Lavanta",
        "description": "Mor ve pembe pastel tonlarda yumuşak bir görünüm.",
        "preview": {"bg": "#f3e8ff", "fg": "#6d28d9"},
    },
}


def user_is_active(user: User | None) -> bool:
    if user is None:
        return False
    return (user.employment_status or "aktif").strip().lower() == "aktif"


def active_users_query(include_inactive: bool = False):
    query = User.query
    if not include_inactive:
        query = query.filter(func.lower(User.employment_status) == "aktif")
    return query


def active_user_by_id(
    user_id: int | None, *, include_inactive: bool = False
) -> User | None:
    if user_id is None:
        return None
    return (
        active_users_query(include_inactive=include_inactive)
        .filter(User.id == user_id)
        .first()
    )


def split_license_name(value: str) -> tuple[str, str]:
    if not value:
        return "", ""
    if " - " in value:
        name, key = value.split(" - ", 1)
        return name.strip(), key.strip()
    return value.strip(), ""


def build_qr_code_url(sku: str) -> str:
    code = sanitize_input_text(sku, max_length=64)
    return (
        f"https://api.qrserver.com/v1/create-qr-code/?size=160x160&data={code}"
        if code
        else ""
    )


def generate_unique_sku(prefix: str) -> str:
    cleaned_prefix = (prefix or "SKU").strip().upper()[:8] or "SKU"
    while True:
        code = f"{cleaned_prefix}-{uuid4().hex[:10].upper()}"
        exists_stock = StockItem.query.filter_by(sku=code).first()
        exists_catalog = ProductCatalogEntry.query.filter_by(sku=code).first()
        if not exists_stock and not exists_catalog:
            return code


def create_app() -> Flask:
    data_dir_env = os.environ.get("DATA_DIR")
    if data_dir_env:
        data_dir = Path(data_dir_env)
    else:
        project_root = Path(__file__).resolve().parent.parent
        data_dir = project_root / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    info_upload_dir = data_dir / "info_uploads"
    info_upload_dir.mkdir(parents=True, exist_ok=True)

    app = Flask(__name__)

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL yapılandırılmamış.")

    app.config.from_mapping(
        SECRET_KEY="stok-admin-secret",
        SQLALCHEMY_DATABASE_URI=database_url,
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
    )
    app.config["DATA_DIR"] = data_dir
    app.config["INFO_UPLOAD_DIR"] = info_upload_dir
    app.permanent_session_lifetime = timedelta(hours=8)

    db.init_app(app)
    app.register_blueprint(personnel_lifecycle_bp)

    with app.app_context():
        db.create_all()
        seed_initial_data()

    @app.before_request
    def enforce_login():
        endpoint = request.endpoint or ""
        if endpoint in {"login", "static", "force_password_change"}:
            return
        if endpoint.startswith("static"):
            return

        user = get_active_user()
        if user is not None:
            if user.must_change_password:
                allowed = {"force_password_change", "logout"}
                if endpoint not in allowed:
                    if request.path.startswith("/api/"):
                        return (
                            jsonify(
                                {
                                    "error": "Devam etmek için lütfen ilk giriş şifrenizi güncelleyin.",
                                }
                            ),
                            403,
                        )
                    if request.method == "GET":
                        next_url = request.full_path or request.path
                        if next_url.endswith("?"):
                            next_url = next_url[:-1]
                        if is_safe_redirect_target(next_url):
                            session["post_password_change_redirect"] = next_url
                    return redirect(url_for("force_password_change"))
            return

        if request.path.startswith("/api/"):
            return jsonify({"error": "Bu işlemi yapmak için oturum açın."}), 401

        next_url = ""
        if request.method == "GET":
            next_url = request.full_path or request.path
            if next_url.endswith("?"):
                next_url = next_url[:-1]
        target = next_url if is_safe_redirect_target(next_url) else None
        return redirect(url_for("login", next=target))

    @app.context_processor
    def inject_profile_preferences() -> dict[str, Any]:
        user = get_active_user()
        sidebar_sections = build_sidebar_sections(user, None, request.endpoint)
        theme_key = "varsayilan"
        if user and user.preferred_theme in THEME_OPTIONS:
            theme_key = user.preferred_theme
        theme_meta = THEME_OPTIONS.get(theme_key, THEME_OPTIONS["varsayilan"])
        return {
            "active_user": user,
            "active_system_role": get_system_role(user),
            "active_theme": theme_key,
            "active_theme_meta": theme_meta,
            "active_theme_class": f"theme-{theme_key}",
            "theme_options": THEME_OPTIONS,
            "system_role_labels": SYSTEM_ROLE_LABELS,
            "is_admin_user": has_system_role(user, "admin"),
            "is_super_admin": has_system_role(user, "superadmin"),
            "sidebar_sections": sidebar_sections,
            "build_breadcrumbs": build_breadcrumbs,
        }

    register_auth_routes(
        app,
        {
            "get_active_user": get_active_user,
            "is_safe_redirect_target": is_safe_redirect_target,
            "active_users_query": active_users_query,
            "User": User,
            "func": func,
            "set_active_user": set_active_user,
            "record_activity": record_activity,
            "current_actor_name": current_actor_name,
            "db": db,
        },
    )
    register_inventory_routes(
        app,
        {
            "load_recent_activity": load_recent_activity,
            "load_dashboard_metrics": load_dashboard_metrics,
            "load_inventory_payload": load_inventory_payload,
            "load_license_payload": load_license_payload,
        },
    )
    register_maintenance_routes(
        app,
        {
            "load_maintenance_payload": load_maintenance_payload,
            "create_maintenance_record": create_maintenance_record,
        },
    )
    register_stock_routes(
        app,
        {
            "get_active_user": get_active_user,
            "has_system_role": has_system_role,
            "load_stock_payload": load_stock_payload,
            "load_scrap_inventory_payload": load_scrap_inventory_payload,
        },
    )

    register_admin_routes(
        app,
        {
            "get_active_user": get_active_user,
            "has_system_role": has_system_role,
            "load_admin_panel_payload": load_admin_panel_payload,
            "SYSTEM_ROLE_LABELS": SYSTEM_ROLE_LABELS,
        },
    )
    register_request_routes(
        app,
        {
            "load_request_groups": load_request_groups,
        },
    )

    register_information_routes(
        app,
    {
        "load_information_payload": load_information_payload,
        "parse_int_or_none": parse_int_or_none,
        "InfoCategory": InfoCategory,
        "save_information_image": save_information_image,
        "InfoEntry": InfoEntry,
        "save_information_file": save_information_file,
        "InfoAttachment": InfoAttachment,
        "db": db,
        "record_activity": record_activity,
        "load_information_entry": load_information_entry,
        "remove_information_image": remove_information_image,
        "remove_information_file": remove_information_file,
    },
    )

    register_profile_routes(
        app,
        {
            "get_active_user": get_active_user,
            "has_system_role": has_system_role,
            "active_users_query": active_users_query,
            "User": User,
            "active_user_by_id": active_user_by_id,
            "set_active_user": set_active_user,
            "parse_int_or_none": parse_int_or_none,
            "THEME_OPTIONS": THEME_OPTIONS,
            "record_activity": record_activity,
            "current_actor_name": current_actor_name,
            "db": db,
            "generate_password_hash": generate_password_hash,
        },
    )

    @app.route("/uploads/info/<path:filename>")
    def info_uploads(filename: str):
        upload_dir: Path = app.config["INFO_UPLOAD_DIR"]
        extension = Path(filename).suffix.lower()
        if extension in DISALLOWED_DOWNLOAD_EXTENSIONS:
            abort(404)

        attachment = InfoAttachment.query.filter_by(filename=filename).first()
        content_type = attachment.content_type if attachment else None
        inline_safe = extension in INLINE_SAFE_EXTENSIONS or (
            content_type and content_type.startswith("image/")
        )
        response = send_from_directory(
            upload_dir, filename, as_attachment=not inline_safe
        )
        if content_type:
            response.headers["Content-Type"] = content_type
        return response

           

    def _postgres_connection_parts():
        database_url = current_app.config.get("SQLALCHEMY_DATABASE_URI", "")

        if not database_url.startswith(("postgresql://", "postgresql+psycopg://")):
            raise RuntimeError("PostgreSQL veritabanı yapılandırması bulunamadı.")

        parsed = urlparse(
            database_url.replace("postgresql+psycopg://", "postgresql://", 1)
        )

        if not parsed.hostname or not parsed.path:
            raise RuntimeError("PostgreSQL bağlantı bilgileri geçersiz.")

        return {
            "host": parsed.hostname,
            "port": str(parsed.port or 5432),
            "user": parsed.username or "",
            "password": parsed.password or "",
            "database": parsed.path.lstrip("/"),
        }


    @app.get("/admin-panel/data/export")
    def export_database():
        user = get_active_user()
        if not has_system_role(user, "superadmin"):
            flash(
                "Veri dışa aktarma işlemi için süper admin yetkisi gerekir.", "danger"
            )
            return redirect(url_for("admin_panel", section="data-section"))

        try:
            connection = _postgres_connection_parts()

            timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".dump",
            ) as tmp:
                backup_path = Path(tmp.name)

            env = os.environ.copy()
            env["PGPASSWORD"] = connection["password"]

            result = subprocess.run(
                [
                    "pg_dump",
                    "-Fc",
                    "--no-owner",
                    "--no-privileges",
                    "-h",
                    connection["host"],
                    "-p",
                    connection["port"],
                    "-U",
                    connection["user"],
                    "-d",
                    connection["database"],
                    "-f",
                    str(backup_path),
                ],
                env=env,
                capture_output=True,
                text=True,
                timeout=300,
            )

            if result.returncode != 0:
                current_app.logger.error(
                    "PostgreSQL backup başarısız: %s",
                    result.stderr.strip(),
                )
                backup_path.unlink(missing_ok=True)
                flash("PostgreSQL veritabanı yedeği oluşturulamadı.", "danger")
                return redirect(
                    url_for("admin_panel", section="data-section")
                )

            record_activity(
                area="sistem",
                action="PostgreSQL veritabanı yedeği indirildi",
                description="Sistem yöneticisi PostgreSQL veritabanının tam yedeğini indirdi.",
                actor=current_actor_name(),
            )
            db.session.commit()

            response = send_file(
                backup_path,
                as_attachment=True,
                download_name=f"stok-postgresql-{timestamp}.dump",
                mimetype="application/octet-stream",
            )

            @after_this_request
            def remove_backup_file(response):
                backup_path.unlink(missing_ok=True)
                return response

            return response

        except Exception:
            current_app.logger.exception(
                "PostgreSQL veritabanı dışa aktarılamadı"
            )
            flash("PostgreSQL veritabanı yedeği oluşturulamadı.", "danger")
            return redirect(
                url_for("admin_panel", section="data-section")
            )


    @app.post("/admin-panel/data/import")
    def import_database_backup():
        user = get_active_user()
        if not has_system_role(user, "superadmin"):
            flash(
                "Veri içe aktarma işlemi için süper admin yetkisi gerekir.",
                "danger",
            )
            return redirect(
                url_for("admin_panel", section="data-section")
            )

        file: FileStorage | None = request.files.get("data_file")

        if file is None or not file.filename:
            flash("Lütfen bir PostgreSQL yedeği seçin.", "warning")
            return redirect(
                url_for("admin_panel", section="data-section")
            )

        filename = secure_filename(file.filename)
        extension = Path(filename).suffix.lower()

        if extension != ".dump":
            flash(
                "Yalnızca PostgreSQL .dump yedekleri içe aktarılabilir.",
                "warning",
            )
            return redirect(
                url_for("admin_panel", section="data-section")
            )

        temp_path: Path | None = None

        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".dump",
            ) as tmp:
                file.save(tmp.name)
                temp_path = Path(tmp.name)

            connection = _postgres_connection_parts()

            env = os.environ.copy()
            env["PGPASSWORD"] = connection["password"]

            db.session.remove()
            db.engine.dispose()

            result = subprocess.run(
                [
                    "pg_restore",
                    "--clean",
                    "--if-exists",
                    "--no-owner",
                    "--no-privileges",
                    "-h",
                    connection["host"],
                    "-p",
                    connection["port"],
                    "-U",
                    connection["user"],
                    "-d",
                    connection["database"],
                    str(temp_path),
                ],
                env=env,
                capture_output=True,
                text=True,
                timeout=600,
            )

            if result.returncode != 0:
                current_app.logger.error(
                    "PostgreSQL restore başarısız: %s",
                    result.stderr.strip(),
                )
                flash(
                    "PostgreSQL veritabanı yedeği geri yüklenemedi.",
                    "danger",
                )
                return redirect(
                    url_for("admin_panel", section="data-section")
                )

            db.session.remove()

            record_activity(
                area="sistem",
                action="PostgreSQL veritabanı yedeği geri yüklendi",
                description="Sistem yöneticisi PostgreSQL veritabanını bir yedekten geri yükledi.",
                actor=current_actor_name(),
            )
            db.session.commit()

            flash(
                "PostgreSQL veritabanı yedeği başarıyla geri yüklendi.",
                "success",
            )

        except subprocess.TimeoutExpired:
            current_app.logger.exception(
                "PostgreSQL restore zaman aşımına uğradı"
            )
            flash(
                "PostgreSQL geri yükleme işlemi zaman aşımına uğradı.",
                "danger",
            )

        except Exception:
            current_app.logger.exception(
                "PostgreSQL veritabanı içe aktarılamadı"
            )
            flash(
                "PostgreSQL veritabanı yedeği geri yüklenemedi.",
                "danger",
            )

        finally:
            if temp_path:
                temp_path.unlink(missing_ok=True)

        return redirect(
            url_for("admin_panel", section="data-section")
        )

    @app.post("/admin-panel/data/stock/import-excel")
    def import_stock_excel():
        user = get_active_user()
        if not has_system_role(user, "admin"):
            flash("Excel içe aktarma işlemi için yönetici yetkisi gerekir.", "danger")
            return redirect(url_for("admin_panel", section="data-section"))

        file: FileStorage | None = request.files.get("excel_file")
        if file is None or not file.filename:
            flash("Lütfen bir Excel dosyası seçin.", "warning")
            return redirect(url_for("admin_panel", section="data-section"))

        filename = secure_filename(file.filename)
        extension = Path(filename).suffix.lower()
        if extension not in {".xlsx", ".xlsm"}:
            flash(
                "Yalnızca .xlsx uzantılı Excel dosyaları içe aktarılabilir.", "warning"
            )
            return redirect(url_for("admin_panel", section="data-section"))

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=extension) as tmp:
                file.save(tmp.name)
                temp_path = Path(tmp.name)
        except Exception:
            flash("Yüklenen dosya kaydedilirken bir hata oluştu.", "danger")
            return redirect(url_for("admin_panel", section="data-section"))

        try:
            workbook = load_workbook(temp_path, data_only=True)
            sheet = workbook.active
        except Exception:
            temp_path.unlink(missing_ok=True)
            flash("Geçerli bir Excel dosyası okunamadı.", "danger")
            return redirect(url_for("admin_panel", section="data-section"))

        try:
            header_map = {
                "başlık": "title",
                "kategori": "category",
                "donanım tipi": "hardware_type",
                "marka": "brand",
                "model": "model",
                "miktar": "quantity",
                "birim": "unit",
                "durum": "status",
                "kaynak": "source_type",
                "referans": "reference_code",
                "not": "note",
                "seri no": "serial_no",
                "garanti bitiş": "warranty_end_date",
                "fabrika": "factory",
                "departman": "department",
                "sorumlu": "responsible",
            }

            column_indexes: dict[str, int] = {}
            header_row = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True), []
            )
            for idx, cell in enumerate(header_row or []):
                header_value = str(cell or "").strip().lower()
                if header_value in header_map:
                    column_indexes[header_map[header_value]] = idx

            if "title" not in column_indexes:
                flash(
                    "Tablonun ilk satırında en azından 'Başlık' sütunu bulunmalı.",
                    "warning",
                )
                return redirect(url_for("admin_panel", section="data-section"))

            category_lookup = {
                **{key: key for key in STOCK_CATEGORY_LABELS.keys()},
                **{label.lower(): key for key, label in STOCK_CATEGORY_LABELS.items()},
            }
            status_lookup = {
                **{key: key for key in STOCK_STATUS_LABELS.keys()},
                **{label.lower(): key for key, label in STOCK_STATUS_LABELS.items()},
            }
            source_lookup = {
                **{key: key for key in STOCK_SOURCE_LABELS.keys()},
                **{label.lower(): key for key, label in STOCK_SOURCE_LABELS.items()},
            }

            def normalize_lookup(
                raw_value: Any, lookup: dict[str, str], fallback: str
            ) -> str:
                if raw_value is None:
                    return fallback
                normalized_value = str(raw_value).strip().lower()
                return lookup.get(normalized_value, fallback)

            imported_count = 0
            current_actor = current_actor_name()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                values = list(row or [])
                if not any(values):
                    continue

                def pick(key: str) -> Any:
                    index = column_indexes.get(key)
                    return (
                        values[index]
                        if index is not None and index < len(values)
                        else None
                    )

                title = str(pick("title") or "").strip()
                if not title:
                    continue

                category_value = normalize_lookup(
                    pick("category"), category_lookup, fallback="envanter"
                )
                status_value = normalize_lookup(
                    pick("status"), status_lookup, fallback="stokta"
                )
                source_value = normalize_lookup(
                    pick("source_type"), source_lookup, fallback="manual"
                )
                quantity = parse_int_or_none(pick("quantity")) or 1
                unit_value = str(pick("unit") or "").strip() or None
                reference_code = str(pick("reference_code") or "").strip() or None
                note_value = str(pick("note") or "").strip() or None
                serial_no = str(pick("serial_no") or "").strip() or None
                warranty_end_date = parse_excel_date(pick("warranty_end_date"))

                metadata = {
                    "hardware_type": str(pick("hardware_type") or "").strip() or None,
                    "brand": str(pick("brand") or "").strip() or None,
                    "model": str(pick("model") or "").strip() or None,
                    "factory": str(pick("factory") or "").strip() or None,
                    "department": str(pick("department") or "").strip() or None,
                    "responsible": str(pick("responsible") or "").strip() or None,
                }

                stock_item = StockItem(
                    source_type=source_value,
                    title=title,
                    category=category_value,
                    quantity=quantity,
                    unit=unit_value,
                    status=status_value,
                    reference_code=reference_code,
                    note=note_value,
                    sku=generate_unique_sku("STK"),
                    serial_no=serial_no,
                    warranty_end_date=warranty_end_date,
                )
                stock_item.metadata_payload = {k: v for k, v in metadata.items() if v}
                db.session.add(stock_item)
                db.session.flush()

                record_stock_log(
                    stock_item,
                    "Excel içe aktarım",
                    action_type="in",
                    performed_by=current_actor,
                    quantity_change=quantity,
                    note=note_value,
                )
                record_stock_audit(
                    stock_item,
                    old_quantity=0,
                    new_quantity=quantity,
                    performed_by=current_actor,
                )

                imported_count += 1

            db.session.commit()

            record_activity(
                area="sistem",
                action="Excel'den stok içe aktarıldı",
                description=f"Excel içe aktarımıyla {imported_count} stok satırı eklendi.",
                actor=current_actor,
            )
            db.session.commit()

            if imported_count:
                flash(f"Excel'den {imported_count} stok kaydı eklendi.", "success")
            else:
                flash("Excel dosyasında eklenecek satır bulunamadı.", "info")

            return redirect(url_for("admin_panel", section="data-section"))
        finally:
            temp_path.unlink(missing_ok=True)

    @app.post("/admin-panel/data/reset")
    def reset_database_view():
        user = get_active_user()
        if not has_system_role(user, "superadmin"):
            flash("Veritabanını sıfırlamak için süper admin yetkisi gerekir.", "danger")
            return redirect(url_for("admin_panel", section="data-section"))

        data_dir = Path(
            current_app.config.get(
                "DATA_DIR", Path(__file__).resolve().parent.parent / "data"
            )
        )
        info_upload_dir = Path(
            current_app.config.get("INFO_UPLOAD_DIR", data_dir / "info_uploads")
        )

        try:
            db.session.remove()
            db.drop_all()
            db.create_all()
            if info_upload_dir.exists():
                shutil.rmtree(info_upload_dir, ignore_errors=True)
            info_upload_dir.mkdir(parents=True, exist_ok=True)
            seed_initial_data()
            record_activity(
                area="sistem",
                action="Veritabanı sıfırlandı",
                description="Sistem varsayılan başlangıç verileriyle yeniden oluşturuldu.",
                actor=current_actor_name(),
            )
            db.session.commit()
        except Exception:  # pragma: no cover - güvenlik amaçlı kayıt
            current_app.logger.exception("Veritabanı sıfırlanamadı")
            flash("Veritabanı sıfırlanırken bir hata oluştu.", "danger")
            return redirect(url_for("admin_panel", section="data-section"))

        flash("Veritabanı varsayılan verilerle yeniden oluşturuldu.", "success")
        return redirect(url_for("admin_panel", section="data-section"))

    @app.post("/admin-panel/users")
    def create_user():
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            flash(
                "Yeni kullanıcı oluşturmak için süper admin yetkisi gerekir.", "danger"
            )
            return redirect(url_for("admin_panel"))

        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()
        email = (request.form.get("email") or "").strip()
        system_role = (
            request.form.get("system_role") or "user"
        ).strip().lower() or "user"

        if system_role not in {"user", "admin"}:
            system_role = "user"

        if not all([username, first_name, last_name, email]):
            flash("Lütfen tüm alanları doldurun.", "danger")
            return redirect(url_for("admin_panel"))

        if len(password) < 8:
            flash("Şifre en az 8 karakter olmalıdır.", "warning")
            return redirect(url_for("admin_panel"))

        existing_username = User.query.filter_by(username=username).first()
        existing_email = User.query.filter_by(email=email).first()
        if existing_username or existing_email:
            flash("Bu kullanıcı adı veya e-posta zaten kullanılıyor.", "warning")
            return redirect(url_for("admin_panel"))

        user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            role="",  # roller ileride admin panelinden düzenlenecek
            department="",
            password_hash=generate_password_hash(password),
            system_role=system_role,
            must_change_password=True,
        )
        db.session.add(user)
        db.session.flush()

        record_activity(
            area="kullanici",
            action="Kullanıcı oluşturuldu",
            description=f"{first_name} {last_name} ({username}) eklendi.",
            actor=current_actor_name(),
            metadata={
                "user_id": user.id,
                "email": email,
                "system_role": system_role,
            },
        )

        db.session.commit()

        flash("Yeni kullanıcı başarıyla oluşturuldu.", "success")
        return redirect(url_for("admin_panel"))

    def _user_assignment_counts(user: User) -> dict[str, int]:
        person_name = f"{user.first_name} {user.last_name}".strip()
        inventory_count = InventoryItem.query.filter(
            InventoryItem.responsible_user_id == user.id
        ).count()
        stock_assignment_count = 0
        if person_name:
            stock_assignment_count = StockAssignment.query.filter(
                func.lower(StockAssignment.assigned_to) == person_name.lower()
            ).count()
        return {
            "inventory_count": inventory_count,
            "stock_assignment_count": stock_assignment_count,
            "total": inventory_count + stock_assignment_count,
        }

    @app.post("/admin-panel/users/<int:user_id>/delete")
    def delete_user(user_id: int):
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            return (
                jsonify(
                    json_error("Kullanıcı silmek için süper admin yetkisi gerekir.")
                ),
                403,
            )

        user = active_user_by_id(user_id, include_inactive=True)
        if user is None:
            return jsonify(json_error("Silinmek istenen kullanıcı bulunamadı.")), 404

        link_counts = _user_assignment_counts(user)
        if link_counts["total"] > 0:
            return (
                jsonify(
                    json_error(
                        "Kullanıcıya bağlı zimmet/atama kayıtları var; önce devir yapın."
                    )
                    | {"counts": link_counts}
                ),
                400,
            )

        active_user = get_active_user()
        was_active_user = active_user is not None and active_user.id == user.id

        display_name = f"{user.first_name} {user.last_name}".strip()
        description = (
            f"{display_name} ({user.username}) kullanıcısı silindi."
            if display_name
            else f"{user.username} kullanıcısı silindi."
        )

        metadata = {"user_id": user.id, "email": user.email}

        if user.system_role == "superadmin":
            remaining_superadmins = (
                User.query.filter(func.lower(User.system_role) == "superadmin")
                .filter(User.id != user.id)
                .count()
            )
            if remaining_superadmins == 0:
                return (
                    jsonify(json_error("Son süper admin kullanıcısı silinemez.")),
                    400,
                )

        db.session.delete(user)
        record_activity(
            area="kullanici",
            action="Kullanıcı silindi",
            description=description,
            actor=current_actor_name(),
            metadata=metadata,
        )
        db.session.commit()

        if was_active_user:
            session.clear()

        return jsonify({"message": "Kullanıcı başarıyla silindi."})

    def _run_user_transfer(
        old_user: User,
        target_user: User,
        delegate_user: User | None,
        *,
        new_department: str | None,
        new_factory: Factory | None,
        dry_run: bool,
    ) -> dict[str, Any]:
        old_name = f"{old_user.first_name} {old_user.last_name}".strip()
        target_name = f"{target_user.first_name} {target_user.last_name}".strip()
        delegate_name = (
            f"{delegate_user.first_name} {delegate_user.last_name}".strip()
            if delegate_user
            else target_name
        )

        inventory_items = (
            InventoryItem.query.options(joinedload(InventoryItem.licenses))
            .filter(InventoryItem.responsible_user_id == old_user.id)
            .all()
        )
        license_rows = [
            license for item in inventory_items for license in (item.licenses or [])
        ]
        open_requests = (
            RequestOrder.query.join(RequestOrder.group)
            .filter(RequestGroup.key == "acik")
            .filter(func.lower(RequestOrder.requested_by) == old_name.lower())
            .all()
            if old_name
            else []
        )

        preview = {
            "inventory_count": len(inventory_items),
            "license_count": len(license_rows),
            "open_request_count": len(open_requests),
        }
        if dry_run:
            return {
                "dry_run": True,
                "preview": preview,
                "success": {"inventory": [], "licenses": [], "requests": []},
                "failed": {"inventory": [], "licenses": [], "requests": []},
            }

        success = {"inventory": [], "licenses": [], "requests": []}
        failed = {"inventory": [], "licenses": [], "requests": []}

        for item in inventory_items:
            try:
                with db.session.begin_nested():
                    item.responsible_user_id = target_user.id
                    if new_department:
                        item.department = new_department
                    if new_factory:
                        item.factory_id = new_factory.id
                    note = f"{old_name} → {target_name} transferi" + (
                        f" · Vekil: {delegate_name}" if delegate_user else ""
                    )
                    add_inventory_event(item, "Personel transferi", note)
                    db.session.flush()
                    success["inventory"].append(
                        {"id": item.id, "inventory_no": item.inventory_no}
                    )
            except Exception as exc:
                failed["inventory"].append(
                    {
                        "id": item.id,
                        "inventory_no": item.inventory_no,
                        "error": str(exc),
                    }
                )

        for license in license_rows:
            try:
                with db.session.begin_nested():
                    record_activity(
                        area="lisans",
                        action="Lisans transfer edildi",
                        description=f"{license.name} · {old_name} -> {target_name}",
                        actor=current_actor_name(),
                        metadata={"license_id": license.id, "item_id": license.item_id},
                    )
                    db.session.flush()
                    success["licenses"].append({"id": license.id, "name": license.name})
            except Exception as exc:
                failed["licenses"].append(
                    {"id": license.id, "name": license.name, "error": str(exc)}
                )

        for order in open_requests:
            try:
                with db.session.begin_nested():
                    order.requested_by = delegate_name or target_name
                    if new_department:
                        order.department = new_department
                    record_activity(
                        area="talep",
                        action="Açık talep transfer edildi",
                        description=f"{order.order_no} · {old_name} -> {order.requested_by}",
                        actor=current_actor_name(),
                        metadata={"order_id": order.id, "order_no": order.order_no},
                    )
                    db.session.flush()
                    success["requests"].append(
                        {"id": order.id, "order_no": order.order_no}
                    )
            except Exception as exc:
                failed["requests"].append(
                    {"id": order.id, "order_no": order.order_no, "error": str(exc)}
                )

        record_activity(
            area="kullanici",
            action="Personel transferi tamamlandı",
            description=(
                f"{old_user.username} -> {target_user.username} | "
                f"Envanter: {len(success['inventory'])}, Lisans: {len(success['licenses'])}, Talep: {len(success['requests'])}"
            ),
            actor=current_actor_name(),
            metadata={
                "old_user_id": old_user.id,
                "target_user_id": target_user.id,
                "delegate_user_id": delegate_user.id if delegate_user else None,
                "new_department": new_department,
                "new_factory_id": new_factory.id if new_factory else None,
                "failed_counts": {k: len(v) for k, v in failed.items()},
            },
        )

        db.session.commit()
        return {
            "dry_run": False,
            "preview": preview,
            "success": success,
            "failed": failed,
        }

    @app.post("/api/users/<int:user_id>/transfer")
    def transfer_user(user_id: int):
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            return (
                jsonify(json_error("Bu işlem için süper admin yetkisi gerekir.")),
                403,
            )

        old_user = active_user_by_id(user_id, include_inactive=True)
        if old_user is None:
            return jsonify(json_error("Transfer edilecek kullanıcı bulunamadı.")), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify(json_error("Geçersiz istek gövdesi.")), 400

        target_user_id = parse_int_or_none(data.get("target_user_id"))
        delegate_user_id = parse_int_or_none(data.get("delegate_user_id"))
        dry_run = bool(data.get("dry_run", True))
        new_department = sanitize_input_text(data.get("new_department")) or None
        new_factory_id = parse_int_or_none(data.get("new_factory_id"))

        target_user = active_user_by_id(target_user_id)
        delegate_user = (
            active_user_by_id(delegate_user_id) if delegate_user_id else None
        )
        new_factory = Factory.query.get(new_factory_id) if new_factory_id else None

        if target_user is None:
            return jsonify(json_error("Yeni sorumlu kullanıcı geçerli değil.")), 400
        if delegate_user_id and delegate_user is None:
            return jsonify(json_error("Vekil kullanıcı geçerli değil.")), 400
        if new_factory_id and new_factory is None:
            return jsonify(json_error("Yeni fabrika seçimi geçerli değil.")), 400

        report = _run_user_transfer(
            old_user,
            target_user,
            delegate_user,
            new_department=new_department,
            new_factory=new_factory,
            dry_run=dry_run,
        )
        return jsonify(
            {
                "message": (
                    "Transfer önizlemesi hazır."
                    if dry_run
                    else "Transfer işlemi tamamlandı."
                ),
                "report": report,
            }
        )

    @app.post("/admin-panel/users/transfer")
    def transfer_user_assignments():
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            return (
                jsonify(json_error("Bu işlem için süper admin yetkisi gerekir.")),
                403,
            )

        old_user_id = parse_int_or_none(request.form.get("old_user_id"))
        target_user_id = parse_int_or_none(request.form.get("new_user_id"))
        delegate_user_id = parse_int_or_none(request.form.get("delegate_user_id"))
        new_department = sanitize_input_text(request.form.get("new_department")) or None
        new_factory_id = parse_int_or_none(request.form.get("new_factory_id"))

        if old_user_id is None:
            return jsonify(json_error("Eski kullanıcı seçimi zorunludur.")), 400

        old_user = active_user_by_id(old_user_id, include_inactive=True)
        target_user = active_user_by_id(target_user_id)
        delegate_user = (
            active_user_by_id(delegate_user_id) if delegate_user_id else None
        )
        new_factory = Factory.query.get(new_factory_id) if new_factory_id else None

        if old_user is None:
            return jsonify(json_error("Transfer edilecek kullanıcı bulunamadı.")), 404
        if target_user is None:
            return jsonify(json_error("Yeni sorumlu kullanıcı geçerli değil.")), 400
        if delegate_user_id and delegate_user is None:
            return jsonify(json_error("Vekil kullanıcı geçerli değil.")), 400
        if new_factory_id and new_factory is None:
            return jsonify(json_error("Yeni fabrika seçimi geçerli değil.")), 400

        report = _run_user_transfer(
            old_user,
            target_user,
            delegate_user,
            new_department=new_department,
            new_factory=new_factory,
            dry_run=False,
        )
        return jsonify({"message": "Transfer işlemi tamamlandı.", "report": report})

    @app.post("/admin-panel/users/<int:user_id>/deactivate")
    def deactivate_user(user_id: int):
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            flash("Bu işlem için süper admin yetkisi gerekir.", "danger")
            return redirect(url_for("admin_panel"))

        user = active_user_by_id(user_id, include_inactive=True)
        if user is None:
            flash("Kullanıcı bulunamadı.", "danger")
            return redirect(url_for("admin_panel"))

        if (user.employment_status or "aktif").lower() == "pasif":
            flash("Kullanıcı zaten pasif durumda.", "info")
            return redirect(url_for("admin_panel"))

        if user.system_role == "superadmin":
            remaining_superadmins = (
                active_users_query(include_inactive=True)
                .filter(func.lower(User.system_role) == "superadmin")
                .filter(func.lower(User.employment_status) == "aktif")
                .filter(User.id != user.id)
                .count()
            )
            if remaining_superadmins == 0:
                flash("Son aktif süper admin pasife alınamaz.", "warning")
                return redirect(url_for("admin_panel"))

        user.employment_status = "pasif"
        user.termination_note = sanitize_input_text(
            request.form.get("termination_note"), max_length=512
        )
        user.termination_date = date.today()

        record_activity(
            area="kullanici",
            action="Kullanici pasife alindi",
            description=f"{user.username} pasife alındı.",
            actor=current_actor_name(),
            metadata={"user_id": user.id, "employment_status": "pasif"},
        )
        db.session.commit()

        if active_user and active_user.id == user.id:
            session.clear()
            flash("Hesabınız pasife alındı.", "warning")
            return redirect(url_for("login"))

        flash("Kullanıcı pasife alındı.", "success")
        return redirect(url_for("admin_panel"))

    @app.post("/admin-panel/users/<int:user_id>/reactivate")
    def reactivate_user(user_id: int):
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            flash("Bu işlem için süper admin yetkisi gerekir.", "danger")
            return redirect(url_for("admin_panel"))

        user = active_user_by_id(user_id, include_inactive=True)
        if user is None:
            flash("Kullanıcı bulunamadı.", "danger")
            return redirect(url_for("admin_panel"))

        if (user.employment_status or "aktif").lower() == "aktif":
            flash("Kullanıcı zaten aktif durumda.", "info")
            return redirect(url_for("admin_panel"))

        user.employment_status = "aktif"
        user.termination_note = None
        user.termination_date = None

        record_activity(
            area="kullanici",
            action="Kullanici aktive edildi",
            description=f"{user.username} yeniden aktifleştirildi.",
            actor=current_actor_name(),
            metadata={"user_id": user.id, "employment_status": "aktif"},
        )
        db.session.commit()

        flash("Kullanıcı yeniden aktifleştirildi.", "success")
        return redirect(url_for("admin_panel"))

    @app.post("/admin-panel/users/<int:user_id>/role")
    def update_user_role(user_id: int):
        active_user = get_active_user()
        if not has_system_role(active_user, "superadmin"):
            return (
                jsonify(
                    json_error("Bu işlemi yapmak için süper admin yetkisi gerekir.")
                ),
                403,
            )

        user = active_user_by_id(user_id, include_inactive=True)
        if user is None:
            return jsonify(json_error("Kullanıcı bulunamadı.")), 404

        data = request.get_json(silent=True) or {}
        new_role = (data.get("system_role") or "").strip().lower()
        if new_role not in SYSTEM_ROLE_LEVELS:
            return jsonify(json_error("Geçersiz yetki değeri.")), 400

        if user.system_role == "superadmin" and new_role != "superadmin":
            remaining = (
                User.query.filter(func.lower(User.system_role) == "superadmin")
                .filter(User.id != user.id)
                .count()
            )
            if remaining == 0:
                return jsonify(json_error("Son süper admin yetkisi düşürülemez.")), 400
            if active_user and active_user.id == user.id:
                return (
                    jsonify(
                        json_error("Aktif kullanıcının yetkisi buradan düşürülemez.")
                    ),
                    400,
                )

        if user.system_role == new_role:
            return jsonify(
                {
                    "user": user.to_dict(),
                    "message": "Yetki güncellendi.",
                }
            )

        user.system_role = new_role
        record_activity(
            area="kullanici",
            action="Yetki güncellendi",
            description=f"{user.username} → {SYSTEM_ROLE_LABELS.get(new_role, new_role)}",
            actor=current_actor_name(),
            metadata={"user_id": user.id, "system_role": new_role},
        )
        db.session.commit()

        return (
            jsonify(
                {
                    "user": user.to_dict(),
                    "message": "Kullanıcı yetkisi güncellendi.",
                }
            ),
            200,
        )

    @app.post("/api/options/<string:option_key>")
    def create_option(option_key: str):
        if option_key == "brands":
            return create_brand()

        model = OPTION_MODEL_MAPPING.get(option_key)
        if not model:
            abort(404)

        try:
            name = parse_option_name(request.get_json())
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if find_existing_by_name(model, name):
            return jsonify({"error": "Bu kayıt zaten mevcut."}), 409

        option = model(name=name)
        db.session.add(option)
        db.session.commit()

        return jsonify(option.to_dict()), 201

    @app.delete("/api/options/<string:option_key>/<int:option_id>")
    def delete_option(option_key: str, option_id: int):
        if option_key == "brands":
            return delete_brand(option_id)

        model = OPTION_MODEL_MAPPING.get(option_key)
        if not model:
            abort(404)

        option = model.query.get(option_id)
        if option is None:
            return jsonify({"error": "Kayıt bulunamadı."}), 404

        db.session.delete(option)
        db.session.commit()
        return ("", 204)

    @app.post("/api/ldap-profiles")
    def create_ldap_profile():
        if not has_system_role(get_active_user(), "admin"):
            return jsonify(json_error("Bu işlemi yapmak için yetkiniz yok.")), 403

        try:
            payload = parse_ldap_profile_payload(request.get_json(silent=True))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        existing = LdapProfile.query.filter(
            func.lower(LdapProfile.name) == payload["name"].lower()
        ).first()
        if existing:
            return jsonify({"error": "Bu profil adı zaten kayıtlı."}), 409

        profile = LdapProfile(**payload)
        db.session.add(profile)
        record_activity(
            area="entegrasyon",
            action="LDAP profili oluşturuldu",
            description=profile.name,
            actor=current_actor_name(),
            metadata={"profile_id": profile.id},
        )
        db.session.commit()
        return (
            jsonify(
                {
                    "profile": profile.to_dict(),
                    "message": "LDAP profili kaydedildi.",
                }
            ),
            201,
        )

    @app.put("/api/ldap-profiles/<int:profile_id>")
    def update_ldap_profile(profile_id: int):
        if not has_system_role(get_active_user(), "admin"):
            return jsonify(json_error("Bu işlemi yapmak için yetkiniz yok.")), 403

        profile = LdapProfile.query.get(profile_id)
        if profile is None:
            return jsonify(json_error("LDAP profili bulunamadı.")), 404

        try:
            payload = parse_ldap_profile_payload(request.get_json(silent=True))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        duplicate = (
            LdapProfile.query.filter(
                func.lower(LdapProfile.name) == payload["name"].lower()
            )
            .filter(LdapProfile.id != profile.id)
            .first()
        )
        if duplicate:
            return jsonify({"error": "Bu profil adı zaten kayıtlı."}), 409

        profile.name = payload["name"]
        profile.host = payload["host"]
        profile.port = payload["port"]
        profile.base_dn = payload["base_dn"]
        profile.bind_dn = payload["bind_dn"]

        record_activity(
            area="entegrasyon",
            action="LDAP profili güncellendi",
            description=profile.name,
            actor=current_actor_name(),
            metadata={"profile_id": profile.id},
        )
        db.session.commit()
        return jsonify(
            {"profile": profile.to_dict(), "message": "LDAP profili güncellendi."}
        )

    @app.delete("/api/ldap-profiles/<int:profile_id>")
    def delete_ldap_profile(profile_id: int):
        if not has_system_role(get_active_user(), "admin"):
            return jsonify(json_error("Bu işlemi yapmak için yetkiniz yok.")), 403

        profile = LdapProfile.query.get(profile_id)
        if profile is None:
            return jsonify(json_error("LDAP profili bulunamadı.")), 404

        db.session.delete(profile)
        record_activity(
            area="entegrasyon",
            action="LDAP profili silindi",
            description=profile.name,
            actor=current_actor_name(),
            metadata={"profile_id": profile.id},
        )
        db.session.commit()
        return ("", 204)

    @app.post("/api/options/brands/<int:brand_id>/models")
    def create_model(brand_id: int):
        brand = Brand.query.get(brand_id)
        if brand is None:
            return jsonify({"error": "Marka bulunamadı."}), 404

        try:
            name = parse_option_name(request.get_json())
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        existing = (
            HardwareModel.query.filter_by(brand_id=brand.id)
            .filter(func.lower(HardwareModel.name) == name.lower())
            .first()
        )
        if existing:
            return jsonify({"error": "Bu model zaten mevcut."}), 409

        model = HardwareModel(name=name, brand=brand)
        db.session.add(model)
        db.session.commit()
        return jsonify(model.to_dict()), 201

    @app.delete("/api/options/models/<int:model_id>")
    def delete_model(model_id: int):
        model = HardwareModel.query.get(model_id)
        if model is None:
            return jsonify({"error": "Model bulunamadı."}), 404

        db.session.delete(model)
        db.session.commit()
        return ("", 204)

    @app.post("/api/inventory")
    def create_inventory():
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        inventory_no = (data.get("inventory_no") or "").strip()
        if not inventory_no:
            return json_error("Envanter numarası zorunludur."), 400

        existing = InventoryItem.query.filter_by(inventory_no=inventory_no).first()
        if existing:
            return json_error("Bu envanter numarası zaten kullanılıyor."), 409

        factory_id = parse_int_or_none(data.get("factory_id"))
        hardware_type_id = parse_int_or_none(data.get("hardware_type_id"))
        brand_id = parse_int_or_none(data.get("brand_id"))
        model_id = parse_int_or_none(data.get("model_id"))
        responsible_user_id = parse_int_or_none(data.get("responsible_user_id"))

        factory = Factory.query.get(factory_id) if factory_id else None
        hardware_type = (
            HardwareType.query.get(hardware_type_id) if hardware_type_id else None
        )
        brand = Brand.query.get(brand_id) if brand_id else None
        model = HardwareModel.query.get(model_id) if model_id else None
        responsible_user = (
            active_user_by_id(responsible_user_id) if responsible_user_id else None
        )

        if not factory:
            return json_error("Geçerli bir fabrika seçin."), 400
        if not hardware_type:
            return json_error("Geçerli bir donanım tipi seçin."), 400
        if not brand:
            return json_error("Geçerli bir marka seçin."), 400
        if not model:
            return json_error("Geçerli bir model seçin."), 400
        if responsible_user_id and not responsible_user:
            return json_error("Geçerli bir kullanıcı seçin."), 400

        department = sanitize_input_text(data.get("department"))
        if not department:
            return json_error("Departman alanı zorunludur."), 400

        item = InventoryItem(
            inventory_no=inventory_no,
            computer_name=(data.get("computer_name") or "").strip() or None,
            factory_id=factory_id,
            department=department,
            hardware_type_id=hardware_type_id,
            responsible_user_id=responsible_user_id,
            brand_id=brand_id,
            model_id=model_id,
            serial_no=(data.get("serial_no") or "").strip() or None,
            ifs_no=(data.get("ifs_no") or "").strip() or None,
            related_machine_no=(data.get("related_machine_no") or "").strip() or None,
            note=(data.get("note") or "").strip() or None,
        )
        db.session.add(item)
        db.session.flush()
        add_inventory_event(item, "Envanter oluşturuldu")
        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return (
            jsonify({"item": serialize_inventory_item(fresh_item)}),
            201,
        )

    @app.patch("/api/inventory/<int:item_id>")
    def update_inventory(item_id: int):
        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        inventory_no = (data.get("inventory_no") or item.inventory_no or "").strip()
        if not inventory_no:
            return json_error("Envanter numarası zorunludur."), 400

        if (
            inventory_no != item.inventory_no
            and InventoryItem.query.filter_by(inventory_no=inventory_no).first()
        ):
            return json_error("Bu envanter numarası zaten kullanılıyor."), 409

        factory_id = parse_int_or_none(data.get("factory_id"))
        hardware_type_id = parse_int_or_none(data.get("hardware_type_id"))
        brand_id = parse_int_or_none(data.get("brand_id"))
        model_id = parse_int_or_none(data.get("model_id"))
        responsible_user_id = parse_int_or_none(data.get("responsible_user_id"))

        factory = Factory.query.get(factory_id) if factory_id else None
        hardware_type = (
            HardwareType.query.get(hardware_type_id) if hardware_type_id else None
        )
        brand = Brand.query.get(brand_id) if brand_id else None
        model = HardwareModel.query.get(model_id) if model_id else None
        responsible_user = (
            active_user_by_id(responsible_user_id) if responsible_user_id else None
        )

        if not factory:
            return json_error("Geçerli bir fabrika seçin."), 400
        if not hardware_type:
            return json_error("Geçerli bir donanım tipi seçin."), 400
        if not brand:
            return json_error("Geçerli bir marka seçin."), 400
        if not model:
            return json_error("Geçerli bir model seçin."), 400
        if responsible_user_id and not responsible_user:
            return json_error("Geçerli bir kullanıcı seçin."), 400

        department = sanitize_input_text(data.get("department"))
        if not department:
            return json_error("Departman alanı zorunludur."), 400

        status = (data.get("status") or item.status or "aktif").strip().lower()
        if status not in INVENTORY_STATUSES:
            return json_error("Geçersiz durum değeri."), 400

        item.inventory_no = inventory_no
        item.computer_name = (data.get("computer_name") or "").strip() or None
        item.factory = factory
        item.department = department
        item.hardware_type = hardware_type
        item.responsible_user = responsible_user
        item.brand = brand
        item.model = model
        item.serial_no = (data.get("serial_no") or "").strip() or None
        item.ifs_no = (data.get("ifs_no") or "").strip() or None
        if "related_machine_no" in data:
            item.related_machine_no = (
                data.get("related_machine_no") or ""
            ).strip() or None
        if "machine_no" in data:
            item.machine_no = (data.get("machine_no") or "").strip() or None
        item.note = (data.get("note") or "").strip() or None
        item.status = status

        add_inventory_event(item, "Envanter bilgileri güncellendi")
        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return jsonify({"item": serialize_inventory_item(fresh_item)})

    @app.post("/api/inventory/<int:item_id>/assign")
    def assign_inventory(item_id: int):
        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        factory_id = parse_int_or_none(data.get("factory_id"))
        responsible_user_id = parse_int_or_none(data.get("responsible_user_id"))
        department = sanitize_input_text(data.get("department"))

        factory = Factory.query.get(factory_id) if factory_id else None
        responsible_user = (
            active_user_by_id(responsible_user_id) if responsible_user_id else None
        )

        if not factory:
            return json_error("Geçerli bir fabrika seçin."), 400
        if responsible_user_id and not responsible_user:
            return json_error("Geçerli bir kullanıcı seçin."), 400
        if not department:
            return json_error("Departman alanı zorunludur."), 400

        item.factory = factory
        item.department = department
        item.responsible_user = responsible_user
        if "related_machine_no" in data:
            item.related_machine_no = (
                data.get("related_machine_no") or ""
            ).strip() or None

        note_parts: list[str] = []
        note_parts.append(f"Fabrika: {factory.name}")
        note_parts.append(f"Departman: {department}")
        if responsible_user:
            note_parts.append(
                f"Sorumlu: {responsible_user.first_name} {responsible_user.last_name}"
            )

        add_inventory_event(item, "Atama güncellendi", " • ".join(note_parts))
        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return jsonify({"item": serialize_inventory_item(fresh_item)})

    @app.post("/api/inventory/<int:item_id>/mark-faulty")
    def mark_inventory_faulty(item_id: int):
        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        reason = (data.get("reason") or "").strip()
        location = (data.get("location") or "").strip()
        note_parts = []
        if reason:
            note_parts.append(f"Arıza Nedeni: {reason}")
        if location:
            note_parts.append(f"Gönderildiği Yer: {location}")

        item.status = "arizali"
        add_inventory_event(item, "Arıza bildirimi", " • ".join(note_parts))
        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return jsonify({"item": serialize_inventory_item(fresh_item)})

    @app.post("/api/inventory/<int:item_id>/stock")
    def move_inventory_to_stock(item_id: int):
        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR

        existing_stock = (
            StockItem.query.options(
                joinedload(StockItem.inventory_item).joinedload(
                    InventoryItem.hardware_type
                ),
                joinedload(StockItem.inventory_item).joinedload(InventoryItem.factory),
                joinedload(StockItem.inventory_item).joinedload(InventoryItem.brand),
                joinedload(StockItem.inventory_item).joinedload(InventoryItem.model),
                joinedload(StockItem.logs),
            )
            .filter(StockItem.inventory_item_id == item.id)
            .order_by(StockItem.id.desc())
            .first()
        )

        if existing_stock and existing_stock.status == "stokta":
            return json_error("Bu envanter kaydı zaten stokta."), 409

        category_value = determine_stock_category_from_inventory(item)
        if existing_stock:
            category_value = normalize_stock_category(
                existing_stock.category, fallback=category_value
            )

        item.status = "stokta"
        add_inventory_event(item, "Stok girişi", note, performed_by=actor)

        log_entry = None
        if existing_stock:
            metadata_payload = build_inventory_stock_metadata(item)
            metadata_payload = remove_assignment_only_metadata(
                metadata_payload, category_value
            )
            existing_stock.status = "stokta"
            existing_stock.quantity = 1
            existing_stock.reference_code = item.inventory_no
            existing_stock.source_type = "inventory"
            existing_stock.inventory_item = item
            if note:
                existing_stock.note = note
            existing_stock.metadata_payload = {
                key: value for key, value in metadata_payload.items() if value
            }
            log_entry = record_stock_log(
                existing_stock,
                "Envanter stoğa geri alındı",
                action_type="in",
                performed_by=actor,
                quantity_change=0,
                note=note,
                metadata={"inventory_no": item.inventory_no},
            )
            stock_item = existing_stock
        else:
            stock_item = create_stock_item_from_inventory(
                item,
                note=note,
                actor=actor,
            )

        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        payload: dict[str, Any] = {"item": serialize_inventory_item(fresh_item)}
        if stock_item:
            fresh_stock = get_stock_item_with_relations(stock_item.id)
            if fresh_stock:
                payload["stock_item"] = serialize_stock_item(fresh_stock)
                if log_entry:
                    payload["log"] = serialize_stock_log(log_entry)
                elif fresh_stock.logs:
                    payload["log"] = serialize_stock_log(fresh_stock.logs[0])
        return jsonify(payload)

    @app.post("/api/inventory/<int:item_id>/scrap")
    def scrap_inventory(item_id: int):
        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        item.status = "hurda"
        if note:
            item.note = note
        add_inventory_event(item, "Hurdaya ayırma", note)
        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return jsonify({"item": serialize_inventory_item(fresh_item)})

    @app.post("/api/licenses/<int:license_id>/stock")
    def move_license_to_stock(license_id: int):
        license = (
            InventoryLicense.query.options(
                joinedload(InventoryLicense.item)
                .joinedload(InventoryItem.factory)
                .joinedload(InventoryItem.hardware_type),
                joinedload(InventoryLicense.item).joinedload(InventoryItem.brand),
                joinedload(InventoryLicense.item).joinedload(InventoryItem.model),
                joinedload(InventoryLicense.item).joinedload(
                    InventoryItem.responsible_user
                ),
                joinedload(InventoryLicense.item).joinedload(InventoryItem.events),
            )
            .filter_by(id=license_id)
            .first()
        )
        if license is None:
            return json_error("Lisans kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR

        associated_item = license.item
        stock_item = create_stock_item_from_license(license, note=note, actor=actor)

        license.status = "pasif"
        license.item = None

        if associated_item:
            add_inventory_event(
                associated_item,
                "Lisans stoklandı",
                note or f"{license.name} lisansı stok listesine taşındı.",
                performed_by=actor,
            )

        fresh_license = (
            InventoryLicense.query.options(
                joinedload(InventoryLicense.item).joinedload(
                    InventoryItem.responsible_user
                ),
                joinedload(InventoryLicense.item).joinedload(
                    InventoryItem.hardware_type
                ),
                joinedload(InventoryLicense.item).joinedload(InventoryItem.factory),
                joinedload(InventoryLicense.item).joinedload(InventoryItem.events),
            )
            .filter_by(id=license.id)
            .first()
        )
        response: dict[str, Any] = {
            "message": "Lisans stok listesine taşındı.",
            "license": (
                serialize_license_record(fresh_license) if fresh_license else None
            ),
        }
        fresh_stock = get_stock_item_with_relations(stock_item.id)
        if fresh_stock:
            response["stock_item"] = serialize_stock_item(fresh_stock)
            if fresh_stock.logs:
                response["log"] = serialize_stock_log(fresh_stock.logs[0])
        return jsonify(response)

    @app.post("/api/stock")
    def create_stock_entry():
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        title = sanitize_input_text(data.get("title"))
        if not title:
            return json_error("Stok adı zorunludur."), 400

        category = normalize_stock_category(data.get("category"))
        quantity = parse_int_or_none(data.get("quantity"))
        if quantity is None:
            quantity = 1
        if quantity < 1:
            return json_error("Miktar en az 1 olmalıdır."), 400
        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR
        reference_code = sanitize_input_text(data.get("reference_code")) or None
        unit = sanitize_input_text(data.get("unit"), max_length=32) or None
        serial_no = sanitize_input_text(data.get("serial_no"), max_length=128) or None
        warranty_end_date = parse_excel_date(data.get("warranty_end_date"))

        try:
            metadata_payload = prepare_stock_metadata(
                category,
                sanitize_metadata_payload(data.get("metadata")),
                include_assignment_fields=False,
            )
        except ValueError as exc:
            return json_error(str(exc)), 400

        if not reference_code:
            reference_code = (
                metadata_payload.get("inventory_no")
                or metadata_payload.get("license_key")
                or None
            )

        active_user = get_active_user()
        category_ref = resolve_stock_category(category)
        unit_ref = resolve_stock_unit(unit)
        stock_item = StockItem(
            source_type="manual",
            title=title,
            category=category,
            category_id=category_ref.id if category_ref else None,
            quantity=quantity,
            status="stokta",
            reference_code=reference_code,
            unit=unit,
            unit_id=unit_ref.id if unit_ref else None,
            note=note or None,
            sku=generate_unique_sku("STK"),
            serial_no=serial_no,
            warranty_end_date=warranty_end_date,
        )
        stock_item.metadata_payload = {
            k: v for k, v in metadata_payload.items() if v
        }
        db.session.add(stock_item)
        db.session.flush()

        log_entry = record_stock_log(
            stock_item,
            "Manuel stok girişi",
            action_type="in",
            performed_by=actor,
            quantity_change=stock_item.quantity,
            note=note,
        )
        record_stock_movement(
            stock_item,
            operation_type="giris",
            old_quantity=0,
            new_quantity=stock_item.quantity,
            user=active_user,
        )
        record_stock_audit(
            stock_item,
            old_quantity=0,
            new_quantity=stock_item.quantity,
            performed_by=actor,
        )

        fresh_item = get_stock_item_with_relations(stock_item.id)
        response_payload: dict[str, Any] = {
            "stock_item": serialize_stock_item(fresh_item)
        }
        if log_entry:
            response_payload["log"] = serialize_stock_log(log_entry)
        return jsonify(response_payload), 201

    @app.post("/api/stock/<int:item_id>/assign")
    def assign_stock_item(item_id: int):
        stock_item = get_stock_item_with_relations(item_id)
        if stock_item is None:
            return json_error("Stok kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR
        quantity = parse_int_or_none(data.get("quantity")) or 1
        if quantity < 1:
            return json_error("Miktar en az 1 olmalıdır."), 400
        if quantity > max(1, stock_item.quantity):
            return json_error("Stokta yeterli miktar bulunmuyor."), 400

        existing_note = stock_item.note

        category_name = (
            stock_item.category_ref.name
            if stock_item.category_ref
            else stock_item.category
        )
        category_value = normalize_stock_category(category_name)
        metadata_defaults: dict[str, Any] = {}
        if stock_item.metadata_payload:
            metadata_defaults.update(stock_item.metadata_payload)
        if stock_item.inventory_item:
            metadata_defaults.update(
                {
                    k: v
                    for k, v in build_inventory_stock_metadata(
                        stock_item.inventory_item
                    ).items()
                    if v
                }
            )

        metadata_defaults = remove_assignment_only_metadata(
            metadata_defaults, category_value
        )

        try:
            assignment_metadata = prepare_stock_metadata(
                category_value,
                sanitize_metadata_payload(data.get("metadata")),
                defaults=metadata_defaults,
            )
        except ValueError as exc:
            return json_error(str(exc)), 400

        def sanitize(values: dict[str, Any]) -> dict[str, str]:
            cleaned: dict[str, str] = {}
            for key, raw in values.items():
                if raw is None:
                    continue
                text = str(raw).strip()
                if text:
                    cleaned[key] = text
            return cleaned

        combined_metadata = sanitize(metadata_defaults)
        combined_metadata.update(sanitize(assignment_metadata))

        previous_quantity = max(1, stock_item.quantity)
        remaining_quantity = max(0, previous_quantity - quantity)

        active_user = get_active_user()
        remaining_item_id: int | None = None
        stock_item.quantity = quantity
        stock_item.metadata_payload = combined_metadata or None
        stock_item.status = "devredildi"
        if note:
            stock_item.note = note

        remaining_item: StockItem | None = None
        if remaining_quantity > 0:
            remaining_item = StockItem(
                source_type=stock_item.source_type,
                source_id=stock_item.source_id,
                inventory_item_id=stock_item.inventory_item_id,
                license_id=stock_item.license_id,
                reference_code=stock_item.reference_code,
                title=stock_item.title,
                category=stock_item.category,
                category_id=stock_item.category_id,
                quantity=remaining_quantity,
                unit=stock_item.unit,
                unit_id=stock_item.unit_id,
                status="stokta",
                note=existing_note,
            )
            remaining_item.metadata_payload = metadata_defaults or None
            db.session.add(remaining_item)
            db.session.flush()
            remaining_item_id = remaining_item.id

            if stock_item.inventory_item:
                inventory = stock_item.inventory_item
                inventory.status = "aktif"
                add_inventory_event(
                    inventory,
                    "Stoktan atama yapıldı",
                    note or f"{stock_item.title} stoğa alınan ürün atandı.",
                    performed_by=actor,
                )

            log_entry = record_stock_log(
                stock_item,
                "Stoktan atama yapıldı",
                action_type="out",
                performed_by=actor,
                quantity_change=-quantity,
                note=note,
                metadata=assignment_metadata or None,
            )
            record_stock_movement(
                stock_item,
                operation_type="zimmet",
                old_quantity=previous_quantity,
                new_quantity=stock_item.quantity,
                user=active_user,
            )
            record_stock_audit(
                stock_item,
                old_quantity=previous_quantity,
                new_quantity=stock_item.quantity,
                performed_by=actor,
            )
            receipt_code = generate_unique_sku("ZIM")
            assignment_record = StockAssignment(
                stock_item_id=stock_item.id,
                assigned_to=(stock_item.metadata_payload or {}).get("responsible")
                or "Belirtilmedi",
                assigned_department=(stock_item.metadata_payload or {}).get(
                    "department"
                )
                or None,
                quantity=quantity,
                delivery_note=note or None,
                delivered_by=actor,
                delivered_at=datetime.utcnow(),
                receipt_code=receipt_code,
            )
            db.session.add(assignment_record)
            if remaining_item:
                record_stock_movement(
                    remaining_item,
                    operation_type="iade",
                    old_quantity=0,
                    new_quantity=remaining_item.quantity,
                    user=active_user,
                )

            responsible_name = (stock_item.metadata_payload or {}).get("responsible")
            if responsible_name:
                record_activity(
                    area="kullanici",
                    action="Stok ataması",
                    description=f"{stock_item.title} → {responsible_name}",
                    actor=actor,
                    metadata={
                        "stock_item_id": stock_item.id,
                        "category": category_value,
                        "responsible": responsible_name,
                        "inventory_no": (stock_item.metadata_payload or {}).get(
                            "inventory_no"
                        ),
                    },
                )

        fresh_item = get_stock_item_with_relations(stock_item.id)
        response_payload: dict[str, Any] = {
            "stock_item": serialize_stock_item(fresh_item)
        }
        if remaining_item_id:
            response_payload["remaining_stock_item"] = serialize_stock_item(
                get_stock_item_with_relations(remaining_item_id)
            )
        if log_entry:
            response_payload["log"] = serialize_stock_log(log_entry)
        return jsonify(response_payload)

    @app.post("/api/stock/<int:item_id>/mark-faulty")
    def mark_stock_item_faulty(item_id: int):
        stock_item = get_stock_item_with_relations(item_id)
        if stock_item is None:
            return json_error("Stok kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR

        previous_quantity = stock_item.quantity
        active_user = get_active_user()

        stock_item.status = "arizali"
        if note:
            stock_item.note = note
        stock_item.quantity = 0

        if stock_item.inventory_item:
            inventory = stock_item.inventory_item
            inventory.status = "arizali"
            add_inventory_event(
                inventory,
                "Stok ürünü arızalı",
                note or f"{stock_item.title} stok kaydı arızalı işaretlendi.",
                performed_by=actor,
            )

        log_entry = record_stock_log(
            stock_item,
            "Stok ürünü arızalı işaretlendi",
            action_type="warning",
            performed_by=actor,
            note=note,
        )
        record_stock_movement(
            stock_item,
            operation_type="durum_guncelleme",
            old_quantity=previous_quantity,
            new_quantity=stock_item.quantity,
            user=active_user,
        )
        record_stock_audit(
            stock_item,
            old_quantity=previous_quantity,
            new_quantity=stock_item.quantity,
            performed_by=actor,
        )

        db.session.commit()

        fresh_item = get_stock_item_with_relations(stock_item.id)
        response_payload: dict[str, Any] = {
            "stock_item": serialize_stock_item(fresh_item)
        }
        if log_entry:
            response_payload["log"] = serialize_stock_log(log_entry)
        return jsonify(response_payload)

    @app.post("/api/stock/<int:item_id>/scrap")
    def scrap_stock_item(item_id: int):
        stock_item = get_stock_item_with_relations(item_id)
        if stock_item is None:
            return json_error("Stok kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        note = sanitize_input_text(data.get("note"), max_length=512)
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR

        previous_quantity = stock_item.quantity
        active_user = get_active_user()
        stock_item.status = "hurda"
        if note:
            stock_item.note = note

        if stock_item.inventory_item:
            inventory = stock_item.inventory_item
            inventory.status = "hurda"
            add_inventory_event(
                inventory,
                "Stok ürünü hurdaya ayrıldı",
                note or f"{stock_item.title} stok kaydı hurdaya ayrıldı.",
                performed_by=actor,
            )

        log_entry = record_stock_log(
            stock_item,
            "Stok ürünü hurdaya ayrıldı",
            action_type="out",
            performed_by=actor,
            quantity_change=-max(1, previous_quantity),
            note=note,
        )
        record_stock_movement(
            stock_item,
            operation_type="satis",
            old_quantity=previous_quantity,
            new_quantity=0,
            user=active_user,
        )
        record_stock_audit(
            stock_item,
            old_quantity=previous_quantity,
            new_quantity=0,
            performed_by=actor,
        )

        fresh_item = get_stock_item_with_relations(stock_item.id)
        response_payload: dict[str, Any] = {
            "stock_item": serialize_stock_item(fresh_item)
        }
        if log_entry:
            response_payload["log"] = serialize_stock_log(log_entry)
        return jsonify(response_payload)

    @app.post("/api/inventory/<int:item_id>/restore-from-scrap")
    def restore_inventory_from_scrap(item_id: int):
        if not has_system_role(get_active_user(), "superadmin"):
            return jsonify(json_error("Bu işlemi yapmak için yetkiniz yok.")), 403

        item = get_inventory_item_with_relations(item_id)
        if item is None:
            return json_error("Envanter kaydı bulunamadı."), 404

        if (item.status or "").lower() != "hurda":
            return json_error("Bu kayıt hurda durumunda değil."), 400

        note = (request.get_json(silent=True) or {}).get("note")
        cleaned_note = (note or "").strip()

        item.status = "stokta"
        actor = current_actor_name()
        add_inventory_event(
            item,
            "Hurda kaydı geri alındı",
            cleaned_note or f"{item.inventory_no} kaydı stok durumuna döndürüldü.",
            performed_by=actor,
        )


        db.session.commit()

        fresh_item = get_inventory_item_with_relations(item.id)
        return jsonify({"item": serialize_inventory_item(fresh_item)})

    @app.post("/api/requests")
    def create_request():
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        order_no = (data.get("order_no") or "").strip()
        requested_by = (data.get("requested_by") or "").strip()
        department = sanitize_input_text(data.get("department"))
        active_user = get_active_user()
        requested_by_user = active_user
        group_key = (data.get("group_key") or "acik").strip().lower() or "acik"
        lines_payload = data.get("lines")

        if not order_no:
            return json_error("Sipariş numarası zorunludur."), 400
        if RequestOrder.query.filter_by(order_no=order_no).first():
            return json_error("Bu sipariş numarası zaten kayıtlı."), 409
        if not requested_by_user:
            return json_error("Talep sahibi doğrulanamadı."), 401

        requested_by = (
            f"{requested_by_user.first_name} {requested_by_user.last_name}".strip()
            or requested_by_user.username
        )
        department = requested_by_user.department or department or "Belirtilmedi"

        if not requested_by:
            return json_error("Talep sahibi seçin."), 400
        if not department:
            return json_error("Departman bilgisi zorunludur."), 400
        if not isinstance(lines_payload, list) or not lines_payload:
            return json_error("En az bir talep satırı ekleyin."), 400

        target_group = get_request_group_by_key(group_key) or get_request_group_by_key(
            "acik"
        )
        if target_group is None:
            return json_error("Talep grubu bulunamadı."), 400

        order = RequestOrder(
            order_no=order_no,
            requested_by=requested_by,
            department=department,
            group=target_group,
        )
        db.session.add(order)

        for index, raw_line in enumerate(lines_payload, start=1):
            if not isinstance(raw_line, dict):
                return json_error("Talep satırı formatı geçersiz."), 400
            hardware_type = (raw_line.get("hardware_type") or "").strip()
            brand = (raw_line.get("brand") or "").strip()
            model = (raw_line.get("model") or "").strip()
            quantity = parse_int_or_none(raw_line.get("quantity")) or 0
            note = (raw_line.get("note") or "").strip() or None
            category_value = normalize_stock_category(
                raw_line.get("category"),
                fallback="envanter",
            )

            if not hardware_type:
                return json_error(f"{index}. satır için donanım tipi zorunludur."), 400
            if quantity <= 0:
                return json_error(f"{index}. satır için geçerli bir miktar girin."), 400

            order.lines.append(
                RequestLine(
                    hardware_type=hardware_type,
                    brand=brand,
                    model=model,
                    quantity=quantity,
                    note=note,
                    category=category_value,
                )
            )

        db.session.flush()

        record_activity(
            area="talep",
            action="Yeni talep oluşturuldu",
            description=f"{order_no} numaralı talep {len(order.lines)} satır ile kaydedildi.",
            metadata={
                "order_id": order.id,
                "order_no": order.order_no,
                "department": order.department,
                "requested_by": requested_by,
                "requested_by_id": requested_by_user.id,
                "line_count": len(order.lines),
            },
        )

        db.session.commit()

        fresh_order = get_request_order_with_relations(order.id)
        payload = serialize_request_order(fresh_order)
        return (
            jsonify(
                {
                    "order": payload,
                    "message": f"{payload['order_no']} numaralı talep kaydedildi.",
                }
            ),
            201,
        )

    @app.post("/api/requests/<int:order_id>/actions")
    def update_request_status(order_id: int):
        order = get_request_order_with_relations(order_id)
        if order is None:
            return json_error("Talep kaydı bulunamadı."), 404

        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        action_key = (data.get("action") or "").strip().lower()
        quantity = parse_int_or_none(data.get("quantity"))
        if quantity is None:
            quantity = 1
        requested_quantity = quantity
        note = sanitize_input_text(data.get("note"), max_length=512) or None
        actor = sanitize_input_text(data.get("performed_by")) or DEFAULT_EVENT_ACTOR

        target_line_id = parse_int_or_none(data.get("line_id"))
        if target_line_id:
            target_lines = [line for line in order.lines if line.id == target_line_id]
            if not target_lines:
                return json_error("Talep satırı bulunamadı."), 404
        else:
            target_lines = list(order.lines)

        if action_key not in {"stok", "cancel"}:
            return json_error("Geçersiz işlem tipi."), 400

        total_quantity = sum(line.quantity for line in target_lines)
        if action_key == "stok" and requested_quantity > 1:
            return (
                json_error("Tek seferde en fazla 1 adet stok girişi yapılabilir."),
                400,
            )
        if requested_quantity < 1:
            return json_error("Miktar en az 1 olmalıdır."), 400
        if total_quantity <= 0:
            return json_error("Talep satırları için geçerli miktar bulunamadı."), 400
        if requested_quantity > total_quantity:
            return json_error("Maksimum işlem miktarı aşılamaz."), 400

        processed_quantity = 0
        category_value = None
        validated_metadata: dict[str, str] | None = None
        if action_key == "stok":
            first_line = target_lines[0] if target_lines else None
            category_value = normalize_stock_category(
                data.get("category"),
                fallback=first_line.category if first_line else "envanter",
            )
            metadata_defaults = {}
            if first_line:
                metadata_defaults.update(
                    {
                        "hardware_type": first_line.hardware_type,
                        "brand": first_line.brand,
                        "model": first_line.model,
                    }
                )
            if order.department:
                metadata_defaults.setdefault("department", order.department)
            try:
                validated_metadata = prepare_stock_metadata(
                    category_value,
                    sanitize_metadata_payload(data.get("metadata")),
                    defaults=metadata_defaults,
                    include_assignment_fields=False,
                )
            except ValueError as exc:
                return json_error(str(exc)), 400

        created_stock_items: list[StockItem] = []
        processed_line_snapshots: list[dict[str, Any]] = []

        def capture_snapshot(line: RequestLine, qty: int) -> None:
            if qty <= 0:
                return
            order.snapshots.append(
                RequestLineSnapshot(
                    hardware_type=line.hardware_type,
                    brand=line.brand,
                    model=line.model,
                    quantity=qty,
                    note=line.note or note,
                    category=line.category,
                    action=action_key,
                )
            )
            processed_line_snapshots.append(
                {
                    "hardware_type": line.hardware_type,
                    "brand": line.brand,
                    "model": line.model,
                    "quantity": qty,
                    "note": line.note or note,
                    "category": line.category,
                    "action": action_key,
                }
            )

        if action_key == "stok":
            remaining_quantity = min(requested_quantity, total_quantity)
            lines_to_remove: list[RequestLine] = []
            for line in target_lines:
                if remaining_quantity <= 0:
                    break
                available_quantity = max(0, line.quantity)
                if available_quantity <= 0:
                    continue
                fulfill_quantity = min(available_quantity, remaining_quantity)
                if fulfill_quantity <= 0:
                    continue
                capture_snapshot(line, fulfill_quantity)
                created_stock_items.append(
                    create_stock_item_from_request_line(
                        order,
                        line,
                        quantity=fulfill_quantity,
                        note=note,
                        actor=actor,
                        category=category_value,
                        metadata=validated_metadata,
                    )
                )
                processed_quantity += fulfill_quantity
                remaining_quantity -= fulfill_quantity
                if fulfill_quantity >= available_quantity:
                    lines_to_remove.append(line)
                else:
                    line.quantity = available_quantity - fulfill_quantity

            for line in lines_to_remove:
                if line in order.lines:
                    order.lines.remove(line)
                db.session.delete(line)

            if processed_quantity <= 0:
                return json_error("İşlem yapılacak geçerli miktar bulunamadı."), 400
        else:
            remaining_quantity = min(requested_quantity, total_quantity)
            lines_to_remove: list[RequestLine] = []
            for line in target_lines:
                if remaining_quantity <= 0:
                    break
                available_quantity = max(0, line.quantity)
                if available_quantity <= 0:
                    continue
                cancel_quantity = min(available_quantity, remaining_quantity)
                if cancel_quantity <= 0:
                    continue
                capture_snapshot(line, cancel_quantity)
                processed_quantity += cancel_quantity
                remaining_quantity -= cancel_quantity
                if cancel_quantity >= available_quantity:
                    lines_to_remove.append(line)
                else:
                    line.quantity = available_quantity - cancel_quantity

            for line in lines_to_remove:
                if line in order.lines:
                    order.lines.remove(line)
                db.session.delete(line)

            if processed_quantity <= 0:
                return json_error("İşlem yapılacak geçerli miktar bulunamadı."), 400

        remaining_total = sum(line.quantity for line in order.lines)
        if action_key == "stok":
            if remaining_total <= 0:
                target_group_key = "kapandi"
                action_label = "Talep stok girişiyle kapandı"
            else:
                target_group_key = "acik"
                action_label = "Talep stok işlemi"
        else:
            if remaining_total <= 0:
                target_group_key = "iptal"
                action_label = "Talep iptal edildi"
            else:
                target_group_key = "acik"
                action_label = "Talep satırı iptal edildi"

        target_group = get_request_group_by_key(target_group_key)
        if target_group:
            order.group = target_group

        if target_group_key != "acik" and (order.snapshots or processed_line_snapshots):
            for line in list(order.lines):
                order.lines.remove(line)
                db.session.delete(line)

            snapshot_sources: list[dict[str, Any]] = []
            if order.snapshots:
                snapshot_sources.extend(
                    [
                        {
                            "hardware_type": snapshot.hardware_type,
                            "brand": snapshot.brand,
                            "model": snapshot.model,
                            "quantity": snapshot.quantity,
                            "note": snapshot.note,
                            "category": snapshot.category,
                        }
                        for snapshot in order.snapshots
                    ]
                )
            else:
                snapshot_sources.extend(processed_line_snapshots)

            for snapshot in snapshot_sources:
                archived_line = RequestLine(
                    hardware_type=snapshot["hardware_type"],
                    brand=snapshot["brand"],
                    model=snapshot["model"],
                    quantity=snapshot["quantity"],
                    note=snapshot.get("note"),
                    category=snapshot["category"],
                )
                order.lines.append(archived_line)

        db.session.flush()

        record_activity(
            area="talep",
            action=action_label,
            description=note,
            actor=actor,
            metadata={
                "order_id": order.id,
                "order_no": order.order_no,
                "quantity": requested_quantity,
                "target_group": order.group.key if order.group else None,
            },
        )

        db.session.commit()

        fresh_order = get_request_order_with_relations(order.id)
        payload = serialize_request_order(fresh_order)
        message = f"{payload['order_no']} numaralı talep için işlem kaydedildi."
        response_payload: dict[str, Any] = {"order": payload, "message": message}
        if created_stock_items:
            response_payload["stock_items"] = [
                serialize_stock_item(get_stock_item_with_relations(item.id))
                for item in created_stock_items
                if item
            ]
        return jsonify(response_payload)

    @app.post("/api/catalog/products")
    def create_catalog_product():
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return json_error("Geçersiz JSON gövdesi."), 400

        department = sanitize_input_text(data.get("department"))
        if not department:
            return json_error("Departman alanı zorunludur."), 400

        try:
            usage_area_id = int(data.get("usage_area_id"))
            license_name_id = int(data.get("license_name_id"))
            info_category_id = int(data.get("info_category_id"))
            factory_id = int(data.get("factory_id"))
            hardware_type_id = int(data.get("hardware_type_id"))
            brand_id = int(data.get("brand_id"))
            model_id = int(data.get("model_id"))
        except (TypeError, ValueError):
            return json_error("Lütfen tüm alanları seçin."), 400

        usage_area = UsageArea.query.get(usage_area_id)
        license_name = LicenseName.query.get(license_name_id)
        info_category = InfoCategory.query.get(info_category_id)
        factory = Factory.query.get(factory_id)
        hardware_type = HardwareType.query.get(hardware_type_id)
        brand = Brand.query.get(brand_id)
        model = HardwareModel.query.get(model_id)

        if not all(
            [
                usage_area,
                license_name,
                info_category,
                factory,
                hardware_type,
                brand,
                model,
            ]
        ):
            return json_error("Seçilen kayıtlar doğrulanamadı."), 400

        entry = ProductCatalogEntry(
            sku=generate_unique_sku("PRD"),
            department=department,
            usage_area=usage_area,
            license_name=license_name,
            info_category=info_category,
            factory=factory,
            hardware_type=hardware_type,
            brand=brand,
            model=model,
        )
        db.session.add(entry)
        db.session.flush()

        record_activity(
            area="urun",
            action="Ürün taslağı kaydedildi",
            description=f"{brand.name} {model.name} için taslak oluşturuldu.",
            metadata={
                "entry_id": entry.id,
                "brand": brand.name,
                "model": model.name,
                "factory": factory.name,
                "department": department,
            },
        )

        db.session.commit()

        fresh_entry = (
            ProductCatalogEntry.query.options(
                joinedload(ProductCatalogEntry.usage_area),
                joinedload(ProductCatalogEntry.license_name),
                joinedload(ProductCatalogEntry.info_category),
                joinedload(ProductCatalogEntry.factory),
                joinedload(ProductCatalogEntry.hardware_type),
                joinedload(ProductCatalogEntry.brand),
                joinedload(ProductCatalogEntry.model),
            )
            .filter_by(id=entry.id)
            .first()
        )

        payload = serialize_catalog_entry(fresh_entry)
        return (
            jsonify(
                {
                    "entry": payload,
                    "message": "Ürün taslağı başarıyla kaydedildi.",
                }
            ),
            201,
        )

    @app.delete("/api/catalog/products/<int:entry_id>")
    def delete_catalog_product(entry_id: int):
        entry = (
            ProductCatalogEntry.query.options(
                joinedload(ProductCatalogEntry.brand),
                joinedload(ProductCatalogEntry.model),
            )
            .filter_by(id=entry_id, is_deleted=False)
            .first()
        )
        if entry is None:
            return jsonify({"error": "Kayıt bulunamadı."}), 404

        brand_name = entry.brand.name if entry.brand else ""
        model_name = entry.model.name if entry.model else ""

        entry.is_deleted = True

        record_activity(
            area="urun",
            action="Ürün taslağı silindi",
            description=f"{brand_name} {model_name} taslağı kaldırıldı.",
            metadata={"entry_id": entry_id},
        )

        db.session.commit()
        return ("", 204)

    @app.get("/api/license-names")
    def list_license_names():
        names = [
            license_name.to_dict()
            for license_name in LicenseName.query.order_by(LicenseName.name)
        ]
        return jsonify({"items": names})

    @app.route("/islem-kayitlari")
    def activity_logs():
        if not has_system_role(get_active_user(), "admin"):
            flash("İşlem kayıtlarını görüntülemek için yetkiniz yok.", "danger")
            return redirect(url_for("index"))
        logs = load_activity_logs()
        unique_areas = sorted({log.get("area", "") for log in logs if log.get("area")})
        default_area = (
            "kullanici"
            if any(log.get("area") == "kullanici" for log in logs)
            else "all"
        )
        return render_template(
            "activity_logs.html",
            active_page="activity_logs",
            logs=logs,
            log_areas=unique_areas,
            default_activity_area=default_area,
        )

    return app


def load_inventory_payload() -> dict:
    items = (
        InventoryItem.query.options(
            joinedload(InventoryItem.factory),
            joinedload(InventoryItem.hardware_type),
            joinedload(InventoryItem.brand),
            joinedload(InventoryItem.model),
            joinedload(InventoryItem.responsible_user),
            joinedload(InventoryItem.events),
            joinedload(InventoryItem.licenses),
            joinedload(InventoryItem.maintenances),
        )
        .order_by(InventoryItem.inventory_no)
        .all()
    )

    payload = [serialize_inventory_item(item) for item in items]
    hidden_statuses = {"stokta", "hurda"}
    visible_items = [
        item for item in payload if item.get("status") not in hidden_statuses
    ]
    faulty_count = sum(1 for item in visible_items if item["status"] == "arizali")
    departments_set: set[str] = {
        item["department"] for item in visible_items if item.get("department")
    }

    factories = [factory.to_dict() for factory in Factory.query.order_by(Factory.name)]
    hardware_types = [
        ht.to_dict() for ht in HardwareType.query.order_by(HardwareType.name)
    ]
    brand_models = [
        brand.to_dict(include_models=True)
        for brand in Brand.query.options(joinedload(Brand.models)).order_by(Brand.name)
    ]
    users = [
        {
            "id": user.id,
            "name": f"{user.first_name} {user.last_name}",
            "department": user.department,
        }
        for user in active_users_query().order_by(User.first_name, User.last_name)
    ]
    departments_set.update({user["department"] for user in users if user["department"]})
    departments = sorted(departments_set)

    status_choices = [
        {"value": "aktif", "label": "Aktif"},
        {"value": "beklemede", "label": "Beklemede"},
        {"value": "arizali", "label": "Arızalı"},
        {"value": "hurda", "label": "Hurdaya Ayrıldı"},
    ]

    return {
        "inventory_items": visible_items,
        "inventory_faulty_count": faulty_count,
        "factories": factories,
        "hardware_types": hardware_types,
        "brand_models": brand_models,
        "users": users,
        "departments": departments,
        "status_choices": status_choices,
    }


def create_maintenance_record(item_id: int, data: Any) -> tuple[dict[str, Any], int]:
    item = get_inventory_item_with_relations(item_id)
    if item is None:
        return json_error("Envanter kaydı bulunamadı."), 404
    if not is_computer_hardware_type(
        item.hardware_type.name if item.hardware_type else None
    ):
        return (
            json_error(
                "Bakım kaydı yalnızca bilgisayar envanterleri için oluşturulabilir."
            ),
            400,
        )

    if not isinstance(data, dict):
        return json_error("Geçersiz JSON gövdesi."), 400

    performed_by = (
        sanitize_input_text(data.get("performed_by"), max_length=128)
        or current_actor_name()
    )
    note = sanitize_input_text(data.get("note"), max_length=2000)
    performed_at_value = (data.get("performed_at") or "").strip()
    performed_at = datetime.utcnow()
    if performed_at_value:
        try:
            performed_at = datetime.fromisoformat(performed_at_value)
        except ValueError:
            return json_error("Bakım tarihi geçerli bir tarih olmalıdır."), 400

    maintenance = InventoryMaintenance(
        item=item,
        performed_at=performed_at,
        performed_by=performed_by,
        note=note or None,
    )
    db.session.add(maintenance)
    db.session.flush()

    event_note_parts = [f"Bakım tarihi: {performed_at.strftime('%d.%m.%Y %H:%M')}"]
    if note:
        event_note_parts.append(note)
    add_inventory_event(
        item,
        "Bakım Yapıldı",
        " • ".join(event_note_parts),
        performed_by=performed_by,
    )
    db.session.commit()

    return {"maintenance": serialize_maintenance_record(maintenance)}, 201


def load_maintenance_payload() -> dict[str, Any]:
    items = (
        InventoryItem.query.options(
            joinedload(InventoryItem.factory),
            joinedload(InventoryItem.hardware_type),
            joinedload(InventoryItem.brand),
            joinedload(InventoryItem.model),
            joinedload(InventoryItem.responsible_user),
            joinedload(InventoryItem.maintenances),
        )
        .order_by(InventoryItem.inventory_no)
        .all()
    )

    computers: list[dict[str, Any]] = []
    for item in items:
        if not is_computer_hardware_type(
            item.hardware_type.name if item.hardware_type else None
        ):
            continue
        if (item.status or "").lower() in {"hurda", "stokta"}:
            continue

        maintenances = [
            serialize_maintenance_record(record) for record in item.maintenances
        ]
        last_maintenance = item.maintenances[0] if item.maintenances else None
        maintenance_status_payload = calculate_maintenance_status(
            last_maintenance.performed_at if last_maintenance else None
        )
        maintenance_status = maintenance_status_payload["label"]
        maintenance_status_key = maintenance_status_payload["status"]
        maintenance_status_class = maintenance_status_badge_class(
            maintenance_status_key
        )

        responsible = (
            f"{item.responsible_user.first_name} {item.responsible_user.last_name}"
            if item.responsible_user
            else "Atama bekliyor"
        )
        brand_model = " ".join(
            filter(
                None,
                [
                    item.brand.name if item.brand else "",
                    item.model.name if item.model else "",
                ],
            )
        )
        search_tokens = [
            item.inventory_no,
            item.computer_name,
            responsible,
            item.department,
            brand_model,
            item.hardware_type.name if item.hardware_type else "",
            maintenance_status,
        ]
        computers.append(
            {
                "id": item.id,
                "inventory_no": item.inventory_no,
                "computer_name": item.computer_name or "",
                "responsible": responsible,
                "department": item.department or "",
                "brand_model": brand_model or "-",
                "hardware_type": item.hardware_type.name if item.hardware_type else "",
                "last_maintenance_at": maintenance_status_payload[
                    "last_maintenance_display"
                ],
                "days_since_maintenance": maintenance_status_payload[
                    "days_since_maintenance"
                ],
                "days_until_due": maintenance_status_payload["days_until_due"],
                "maintenance_status": maintenance_status,
                "maintenance_status_key": maintenance_status_key,
                "maintenance_status_class": maintenance_status_class,
                "maintenance_row_class": maintenance_row_class(maintenance_status_key),
                "maintenances": maintenances,
                "search_index": " ".join(
                    token for token in search_tokens if token
                ).lower(),
            }
        )

    return {
        "maintenance_items": computers,
        "maintenance_total_count": len(computers),
        "maintenance_due_count": sum(
            1
            for item in computers
            if item["maintenance_status_key"] in {"overdue", "none", "warning"}
        ),
    }


def lifecycle_flags_payload(status: str | None) -> dict[str, bool]:
    return {
        "is_active": status == "aktif",
        "is_location_changed": status == "yer_degisti",
        "is_exit_pending": status == "ayrilis_bekliyor",
        "is_exited": status == "ayrildi",
    }


def get_person_lifecycle_status(person_name: str | None) -> str | None:
    if not person_name:
        return None
    normalized = quote_plus(person_name.strip().lower())
    if not normalized:
        return None
    activity = (
        ActivityLog.query.filter(
            ActivityLog.area == "personnel_lifecycle",
            ActivityLog.metadata_json.isnot(None),
            ActivityLog.metadata_json["person_key"].as_string() == normalized,
        )
        .order_by(ActivityLog.created_at.desc(), ActivityLog.id.desc())
        .first()
    )
    if not activity or not activity.metadata_payload:
        return None
    return activity.metadata_payload.get("lifecycle_status")


def normalize_stock_category(value: str | None, fallback: str = "envanter") -> str:
    if not value:
        return fallback
    normalized = value.strip().lower()
    return normalized if normalized in STOCK_CATEGORY_LABELS else fallback


def assignment_only_keys(category: str) -> set[str]:
    schema = STOCK_METADATA_FIELDS.get(category, [])
    return {field["key"] for field in schema if field.get("assignment_only")}


def remove_assignment_only_metadata(
    metadata: dict[str, Any] | None, category: str
) -> dict[str, Any]:
    if not metadata:
        return {}
    disallowed = assignment_only_keys(category)
    if not disallowed:
        return dict(metadata)
    return {key: value for key, value in metadata.items() if key not in disallowed}


def determine_stock_category_from_inventory(
    item: InventoryItem | None, fallback: str = "envanter"
) -> str:
    if not item:
        return fallback
    hardware_name = (item.hardware_type.name if item.hardware_type else "") or ""
    if "yazıcı" in hardware_name.lower():
        return "yazici"
    return fallback


def serialize_stock_item(stock_item: StockItem) -> dict[str, Any]:
    item = stock_item.inventory_item
    license_record = stock_item.license
    metadata = stock_item.metadata_payload or {}

    category_value = normalize_stock_category(stock_item.category)
    if category_value == "envanter" and item:
        category_value = determine_stock_category_from_inventory(item, category_value)
    if category_value == "envanter" and stock_item.source_type == "license":
        category_value = "lisans"
    if category_value == "envanter" and stock_item.source_type == "request":
        category_value = "talep"

    status_value = normalize_stock_status(stock_item.status)
    source_type = (stock_item.source_type or "manual").lower()
    source_label = STOCK_SOURCE_LABELS.get(source_type, STOCK_SOURCE_LABELS["manual"])

    created_display = (
        stock_item.created_at.strftime("%d.%m.%Y %H:%M")
        if stock_item.created_at
        else ""
    )
    updated_display = (
        stock_item.updated_at.strftime("%d.%m.%Y %H:%M")
        if stock_item.updated_at
        else created_display
    )

    hardware_type = (
        item.hardware_type.name
        if item and item.hardware_type
        else metadata.get("hardware_type", "")
    )
    brand_name = item.brand.name if item and item.brand else metadata.get("brand", "")
    model_name = item.model.name if item and item.model else metadata.get("model", "")

    search_tokens = [
        stock_item.title,
        stock_item.reference_code,
        STOCK_CATEGORY_LABELS.get(category_value, category_value.capitalize()),
        STOCK_STATUS_LABELS.get(status_value, status_value.capitalize()),
        source_label,
        metadata.get("factory"),
        metadata.get("department"),
        hardware_type,
        brand_name,
        model_name,
        metadata.get("license_key"),
        metadata.get("request_no"),
        metadata.get("responsible"),
    ]
    if item:
        search_tokens.extend(
            [
                item.inventory_no,
                item.department,
                item.factory.name if item.factory else "",
                hardware_type,
                item.serial_no,
                item.ifs_no,
            ]
        )
    if license_record:
        search_tokens.extend([license_record.name, license_record.status])

    allow_operations = status_value == "stokta"

    person_name = metadata.get("responsible") or (
        item.responsible_user.first_name + " " + item.responsible_user.last_name
        if item and item.responsible_user
        else None
    )
    person_key = quote_plus((person_name or "").strip().lower()) if person_name else ""
    lifecycle_status = get_person_lifecycle_status(person_name)

    return {
        "id": stock_item.id,
        "sku": stock_item.sku or "",
        "title": stock_item.title,
        "category": category_value,
        "category_label": STOCK_CATEGORY_LABELS.get(
            category_value, category_value.capitalize()
        ),
        "quantity": stock_item.quantity,
        "unit": (
            stock_item.unit_ref.name
            if stock_item.unit_ref
            else stock_item.unit or metadata.get("unit") or "adet"
        ),
        "reference_code": stock_item.reference_code or "",
        "status": status_value,
        "status_label": STOCK_STATUS_LABELS.get(
            status_value, status_value.capitalize()
        ),
        "status_class": STOCK_STATUS_CLASSES.get(status_value, "status-stock"),
        "source_type": source_type,
        "source_label": source_label,
        "note": stock_item.note or "",
        "metadata": metadata,
        "inventory_id": item.id if item else None,
        "inventory_no": item.inventory_no if item else "",
        "hardware_type": hardware_type,
        "brand": brand_name,
        "model": model_name,
        "license_id": license_record.id if license_record else None,
        "license_name": (
            license_record.name if license_record else metadata.get("license_name")
        ),
        "created_display": created_display,
        "updated_display": updated_display,
        "search_index": " ".join(filter(None, search_tokens)).lower(),
        "person_key": person_key,
        "lifecycle_status": lifecycle_status,
        "lifecycle_flags": lifecycle_flags_payload(lifecycle_status),
        "allow_operations": allow_operations,
        "serial_no": stock_item.serial_no
        or (item.serial_no if item else "")
        or metadata.get("serial_no", ""),
        "warranty_end_date": (
            stock_item.warranty_end_date.isoformat()
            if stock_item.warranty_end_date
            else ""
        ),
        "qr_code_url": build_qr_code_url(stock_item.sku or ""),
        "is_critical": status_value == "stokta" and int(stock_item.quantity or 0) < 5,
    }


def serialize_stock_log(log: StockLog) -> dict[str, Any]:
    item = log.stock_item
    status_value = normalize_stock_status(item.status if item else "stokta")
    combined_metadata: dict[str, Any] = {}
    if item and item.metadata_payload:
        combined_metadata.update(item.metadata_payload)
    log_metadata = log.metadata_payload or {}
    if log_metadata:
        combined_metadata.update(log_metadata)
    unit_label = "adet"
    if item and item.unit:
        unit_label = item.unit
    elif item and (item.metadata_payload or {}).get("unit"):
        unit_label = (item.metadata_payload or {}).get("unit")  # type: ignore[arg-type]
    return {
        "id": log.id,
        "stock_item_id": item.id if item else None,
        "title": item.title if item else "",
        "action": log.action,
        "action_type": log.action_type,
        "performed_by": log.performed_by,
        "quantity_change": log.quantity_change,
        "unit": unit_label,
        "note": log.note or "",
        "status": status_value,
        "status_label": STOCK_STATUS_LABELS.get(
            status_value, status_value.capitalize()
        ),
        "status_class": STOCK_STATUS_CLASSES.get(status_value, "status-stock"),
        "created_display": log.created_at.strftime("%d.%m.%Y %H:%M"),
        "metadata": log_metadata,
        "details_metadata": combined_metadata,
    }


def load_stock_payload() -> dict[str, Any]:
    in_stock_statuses = {"stokta", "arizali"}
    items = (
        StockItem.query.options(
            joinedload(StockItem.inventory_item).joinedload(
                InventoryItem.hardware_type
            ),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.factory),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.brand),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.model),
            joinedload(StockItem.license),
            joinedload(StockItem.category_ref),
            joinedload(StockItem.unit_ref),
            joinedload(StockItem.logs),
        )
        .filter(StockItem.is_deleted.is_(False))
        .order_by(StockItem.created_at.desc())
        .all()
    )

    all_stock_items = [serialize_stock_item(item) for item in items]
    stock_items = [
        item for item in all_stock_items if item.get("status") in in_stock_statuses
    ]
    category_counts = Counter(item["category"] for item in stock_items)
    status_counts = Counter(item["status"] for item in stock_items)
    faulty_count = status_counts.get("arizali", 0)

    assignment_map: dict[str, list[dict[str, Any]]] = {}
    for item in all_stock_items:
        if item.get("status") != "devredildi":
            continue
        responsible = (item.get("metadata") or {}).get("responsible")
        if not responsible:
            continue
        assignment_map.setdefault(responsible, []).append(
            {
                "id": item["id"],
                "title": item["title"],
                "hardware_type": item.get("hardware_type") or item.get("title"),
                "category_label": item.get("category_label"),
                "quantity": item.get("quantity"),
                "status": item.get("status"),
                "status_label": item.get("status_label"),
                "updated_display": item.get("updated_display"),
            }
        )

    user_assignments = [
        {
            "responsible": name,
            "items": sorted(
                entries,
                key=lambda payload: payload.get("updated_display") or "",
                reverse=True,
            ),
        }
        for name, entries in sorted(assignment_map.items())
    ]

    categories = [
        {
            "value": key,
            "label": STOCK_CATEGORY_LABELS[key],
            "count": category_counts.get(key, 0),
        }
        for key in STOCK_CATEGORY_LABELS
    ]

    status_summary = [
        {
            "value": key,
            "label": STOCK_STATUS_LABELS[key],
            "count": status_counts.get(key, 0),
        }
        for key in STOCK_STATUS_LABELS
        if key in in_stock_statuses
    ]

    logs = (
        StockLog.query.options(joinedload(StockLog.stock_item))
        .order_by(StockLog.created_at.desc())
        .limit(40)
        .all()
    )

    support_options = build_stock_support_options()

    assignments = (
        StockAssignment.query.order_by(StockAssignment.created_at.desc())
        .limit(100)
        .all()
    )

    return {
        "stock_items": stock_items,
        "stock_logs": [serialize_stock_log(log) for log in logs],
        "stock_categories": categories,
        "stock_status_summary": status_summary,
        "stock_faulty_count": faulty_count,
        "stock_metadata_config": STOCK_METADATA_FIELDS,
        "stock_support_options": support_options,
        "stock_user_assignments": user_assignments,
        "stock_assignments": [
            {
                "id": assignment.id,
                "stock_item_id": assignment.stock_item_id,
                "assigned_to": assignment.assigned_to,
                "assigned_department": assignment.assigned_department or "",
                "quantity": assignment.quantity,
                "delivery_note": assignment.delivery_note or "",
                "delivered_by": assignment.delivered_by,
                "delivered_at": (
                    assignment.delivered_at.strftime("%d.%m.%Y %H:%M")
                    if assignment.delivered_at
                    else ""
                ),
                "receipt_code": assignment.receipt_code,
            }
            for assignment in assignments
        ],
    }


def normalize_stock_status(value: str | None, fallback: str = "stokta") -> str:
    if not value:
        return fallback
    normalized = value.strip().lower()
    return normalized if normalized in STOCK_STATUS_LABELS else fallback


def load_scrap_inventory_payload() -> dict[str, Any]:
    items = (
        InventoryItem.query.options(
            joinedload(InventoryItem.factory),
            joinedload(InventoryItem.hardware_type),
            joinedload(InventoryItem.brand),
            joinedload(InventoryItem.model),
            joinedload(InventoryItem.responsible_user),
            joinedload(InventoryItem.events),
        )
        .filter(func.lower(InventoryItem.status) == "hurda")
        .order_by(InventoryItem.updated_at.desc(), InventoryItem.inventory_no)
        .all()
    )

    scrap_items = [serialize_inventory_item(item) for item in items]

    return {
        "scrap_items": scrap_items,
        "scrap_count": len(scrap_items),
    }


def load_information_entry(entry_id: int) -> InfoEntry | None:
    return (
        InfoEntry.query.options(
            joinedload(InfoEntry.category),
            joinedload(InfoEntry.attachments),
        )
        .filter_by(id=entry_id)
        .first()
    )


def load_information_payload() -> dict[str, Any]:
    entries = (
        InfoEntry.query.options(joinedload(InfoEntry.category))
        .order_by(InfoEntry.created_at.desc())
        .all()
    )
    categories = [
        category.to_dict()
        for category in InfoCategory.query.order_by(InfoCategory.name)
    ]
    return {
        "info_entries": entries,
        "categories": categories,
        "info_count": len(entries),
    }


def save_information_file(file: FileStorage | None) -> tuple[str, str] | None:
    if file is None or not file.filename:
        return None

    original_name = secure_filename(file.filename)
    if not original_name:
        return None

    extension = Path(original_name).suffix.lower()
    allowed_mimetypes = INFO_ALLOWED_EXTENSIONS.get(extension)
    if not allowed_mimetypes:
        flash(
            "Bu dosya türüne izin verilmiyor. Lütfen yalnızca görsel veya belge yükleyin.",
            "warning",
        )
        return None

    content_type = (file.mimetype or "").lower()
    guessed_type, _ = mimetypes.guess_type(original_name)
    guessed_type = (guessed_type or "").lower()
    if allowed_mimetypes and (
        content_type not in allowed_mimetypes and guessed_type not in allowed_mimetypes
    ):
        flash(
            "Dosya içerik türü doğrulanamadı. Lütfen desteklenen bir belge yükleyin.",
            "warning",
        )
        return None

    length = file.content_length
    if length is None or length <= 0:
        current_position = file.stream.tell()
        file.stream.seek(0, os.SEEK_END)
        length = file.stream.tell()
        file.stream.seek(current_position)

    if length is not None and length > MAX_INFO_UPLOAD_SIZE:
        flash(
            "Dosya boyutu en fazla 10 MB olabilir. Lütfen daha küçük bir dosya seçin.",
            "warning",
        )
        return None

    if file.stream.tell() != 0:
        file.stream.seek(0)

    unique_name = f"{uuid4().hex}{extension}" if extension else uuid4().hex
    upload_dir: Path = current_app.config["INFO_UPLOAD_DIR"]
    target = upload_dir / unique_name
    file.save(target)
    return unique_name, original_name


def save_information_image(file: FileStorage | None) -> str | None:
    saved = save_information_file(file)
    return saved[0] if saved else None


def remove_information_file(filename: str | None) -> None:
    if not filename:
        return

    upload_dir: Path = current_app.config["INFO_UPLOAD_DIR"]
    target = upload_dir / filename
    try:
        target.unlink()
    except FileNotFoundError:
        pass


def remove_information_image(filename: str | None) -> None:
    remove_information_file(filename)


COMPUTER_HARDWARE_KEYWORDS = {
    "bilgisayar",
    "laptop",
    "desktop",
    "pc",
    "notebook",
    "dizustu",
    "dizüstü",
    "masaustu",
    "masaüstü",
}


def is_computer_hardware_type(name: str | None) -> bool:
    normalized = (name or "").strip().lower()
    if not normalized:
        return False

    ascii_normalized = (
        normalized.replace("ı", "i")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ş", "s")
        .replace("ö", "o")
        .replace("ç", "c")
    )
    tokens = {
        token
        for token in ascii_normalized.replace("-", " ").replace("/", " ").split()
        if token
    }
    fuzzy_keywords = COMPUTER_HARDWARE_KEYWORDS - {"pc", "dizüstü", "masaüstü"}
    return (
        any(keyword in ascii_normalized for keyword in fuzzy_keywords) or "pc" in tokens
    )


def format_datetime_display(
    value: datetime | None, *, include_time: bool = True
) -> str:
    if not value:
        return ""
    return value.strftime("%d.%m.%Y %H:%M" if include_time else "%d.%m.%Y")


def serialize_maintenance_record(record: InventoryMaintenance) -> dict[str, Any]:
    return {
        "id": record.id,
        "item_id": record.item_id,
        "performed_by": record.performed_by,
        "performed_at": record.performed_at.isoformat(),
        "performed_at_display": format_datetime_display(record.performed_at),
        "performed_date_display": format_datetime_display(
            record.performed_at, include_time=False
        ),
        "note": record.note or "",
        "created_at_display": format_datetime_display(record.created_at),
    }


def serialize_inventory_item(item: InventoryItem) -> dict[str, Any]:
    responsible = (
        f"{item.responsible_user.first_name} {item.responsible_user.last_name}"
        if item.responsible_user
        else "Henüz atanmamış"
    )
    brand_name = item.brand.name if item.brand else ""
    model_name = item.model.name if item.model else ""
    status_value = (item.status or "aktif").lower()

    history = [
        {
            "id": event.id,
            "event_type": event.event_type,
            "performed_by": event.performed_by,
            "performed_at": event.performed_at.strftime("%d.%m.%Y %H:%M"),
            "note": event.note,
        }
        for event in item.events
    ]

    licenses = [serialize_license_record(license) for license in item.licenses]
    maintenances = [
        serialize_maintenance_record(record) for record in item.maintenances
    ]

    search_tokens = [
        item.inventory_no,
        item.computer_name,
        item.factory.name if item.factory else "",
        item.department,
        item.hardware_type.name if item.hardware_type else "",
        responsible,
        brand_name,
        model_name,
        item.serial_no,
        item.ifs_no,
    ]

    person_key = quote_plus((responsible or "").strip().lower()) if responsible else ""
    lifecycle_status = get_person_lifecycle_status(responsible)

    return {
        "id": item.id,
        "inventory_no": item.inventory_no,
        "computer_name": item.computer_name,
        "factory": item.factory.name if item.factory else "",
        "factory_id": item.factory_id,
        "department": item.department,
        "hardware_type": item.hardware_type.name if item.hardware_type else "",
        "hardware_type_id": item.hardware_type_id,
        "responsible": responsible,
        "responsible_user_id": item.responsible_user_id,
        "brand": brand_name,
        "brand_id": item.brand_id,
        "model": model_name,
        "model_id": item.model_id,
        "serial_no": item.serial_no,
        "ifs_no": item.ifs_no,
        "related_machine_no": item.related_machine_no,
        "machine_no": item.machine_no,
        "ip_address": item.related_machine_no,
        "mac_address": item.machine_no,
        "note": item.note,
        "is_ip_printer": bool(item.related_machine_no or item.machine_no),
        "status": status_value,
        "history": history,
        "licenses": licenses,
        "maintenances": maintenances,
        "search_index": " ".join(filter(None, search_tokens)).lower(),
        "person_key": person_key,
        "lifecycle_status": lifecycle_status,
        "lifecycle_flags": lifecycle_flags_payload(lifecycle_status),
    }


def serialize_license_record(license: InventoryLicense) -> dict[str, Any]:
    item = license.item
    responsible_user = item.responsible_user if item else None
    responsible_name = (
        f"{responsible_user.first_name} {responsible_user.last_name}"
        if responsible_user
        else ""
    )
    email = responsible_user.email if responsible_user else ""
    department = responsible_user.department if responsible_user else ""
    inventory_no = item.inventory_no if item else ""
    computer_name = item.computer_name if item else ""
    hardware_type_name = item.hardware_type.name if item and item.hardware_type else ""
    inventory_label = inventory_no
    if inventory_no:
        if computer_name:
            inventory_label = f"{inventory_no} · {computer_name}"
        elif hardware_type_name:
            inventory_label = f"{inventory_no} · {hardware_type_name}"
    factory_name = item.factory.name if item and item.factory else ""
    ifs_no = item.ifs_no if item else ""
    status_value = (license.status or "aktif").lower()
    status_label = LICENSE_STATUS_LABELS.get(status_value, status_value.capitalize())
    display_name, key = split_license_name(license.name)

    history: list[dict[str, Any]] = []
    if item and item.events:
        for event in sorted(item.events, key=lambda e: e.performed_at, reverse=True):
            history.append(
                {
                    "title": event.event_type,
                    "actor": event.performed_by,
                    "note": event.note or "",
                    "performed_at": event.performed_at.strftime("%d.%m.%Y %H:%M"),
                }
            )

    search_tokens = [
        display_name or license.name,
        key,
        responsible_name,
        email,
        department,
        inventory_no,
        computer_name,
        factory_name,
        status_label,
    ]

    return {
        "id": license.id,
        "display_name": display_name or license.name,
        "key": key,
        "raw_name": license.name,
        "status": status_value,
        "status_label": status_label,
        "responsible_id": responsible_user.id if responsible_user else None,
        "responsible_name": responsible_name or "Atama bekliyor",
        "responsible_department": department,
        "email": email,
        "inventory_id": item.id if item else None,
        "inventory_no": inventory_no,
        "inventory_label": inventory_label or inventory_no,
        "computer_name": computer_name,
        "factory": factory_name,
        "department": item.department if item else "",
        "ifs_no": ifs_no,
        "history": history,
        "search_index": " ".join(token for token in search_tokens if token).lower(),
    }


def load_license_payload() -> dict[str, Any]:
    licenses = (
        InventoryLicense.query.options(
            joinedload(InventoryLicense.item).joinedload(
                InventoryItem.responsible_user
            ),
            joinedload(InventoryLicense.item).joinedload(InventoryItem.hardware_type),
            joinedload(InventoryLicense.item).joinedload(InventoryItem.factory),
            joinedload(InventoryLicense.item).joinedload(InventoryItem.events),
        )
        .order_by(InventoryLicense.id)
        .all()
    )

    license_records = [serialize_license_record(license) for license in licenses]

    users = [
        {
            "id": user.id,
            "name": f"{user.first_name} {user.last_name}",
            "email": user.email,
            "department": user.department or "",
        }
        for user in active_users_query().order_by(User.first_name, User.last_name)
    ]

    inventory_options = [
        {
            "id": item.id,
            "inventory_no": item.inventory_no,
            "label": (
                f"{item.inventory_no} · {item.computer_name}"
                if item.computer_name
                else (
                    f"{item.inventory_no} · {item.hardware_type.name}"
                    if item.hardware_type
                    else item.inventory_no
                )
            ),
            "ifs_no": item.ifs_no or "",
            "department": item.department or "",
        }
        for item in InventoryItem.query.options(
            joinedload(InventoryItem.hardware_type)
        ).order_by(InventoryItem.inventory_no)
        if (item.status or "").lower() != "stokta"
    ]

    status_counts = {
        "total": len(license_records),
        "active": sum(1 for record in license_records if record["status"] == "aktif"),
        "passive": sum(1 for record in license_records if record["status"] == "pasif"),
    }

    return {
        "license_records": license_records,
        "license_users": users,
        "license_inventory_options": inventory_options,
        "license_names": [
            ln.to_dict() for ln in LicenseName.query.order_by(LicenseName.name)
        ],
        "license_status_counts": status_counts,
    }


def serialize_request_order(order: RequestOrder) -> dict[str, Any]:
    opened_display = order.opened_at.strftime("%d.%m.%Y %H:%M")
    lines_payload: list[dict[str, Any]] = []
    group_key = order.group.key if order.group else None
    use_snapshots = group_key in {"kapandi", "iptal"}
    search_tokens = [
        order.order_no,
        order.requested_by,
        order.department,
        opened_display,
    ]

    def build_line_payload(
        source_line: RequestLine | RequestLineSnapshot,
    ) -> dict[str, Any]:
        category_value = normalize_stock_category(
            source_line.category, fallback="envanter"
        )
        line_payload = {
            "id": source_line.id,
            "hardware_type": source_line.hardware_type,
            "brand": source_line.brand,
            "model": source_line.model,
            "quantity": source_line.quantity,
            "note": source_line.note,
            "opened_display": opened_display,
            "category": category_value,
            "category_label": STOCK_CATEGORY_LABELS.get(
                category_value, category_value.capitalize()
            ),
        }
        return line_payload

    source_lines = order.snapshots if use_snapshots and order.snapshots else order.lines

    for line in source_lines:
        line_payload = build_line_payload(line)
        lines_payload.append(line_payload)
        search_tokens.extend(
            [
                line_payload["hardware_type"],
                line_payload["brand"],
                line_payload["model"],
                line_payload["category_label"],
                line_payload.get("note"),
            ]
        )

    person_key = (
        quote_plus((order.requested_by or "").strip().lower())
        if order.requested_by
        else ""
    )
    lifecycle_status = get_person_lifecycle_status(order.requested_by)

    return {
        "id": order.id,
        "order_no": order.order_no,
        "requested_by": order.requested_by,
        "department": order.department,
        "opened_display": opened_display,
        "lines": lines_payload,
        "item_count": len(lines_payload),
        "total_quantity": sum(line["quantity"] for line in lines_payload),
        "search_index": " ".join(token for token in search_tokens if token).lower(),
        "group_key": group_key,
        "person_key": person_key,
        "lifecycle_status": lifecycle_status,
        "lifecycle_flags": lifecycle_flags_payload(lifecycle_status),
    }


def serialize_activity_log(log: ActivityLog) -> dict[str, Any]:
    return {
        "id": log.id,
        "area": log.area,
        "action": log.action,
        "description": log.description,
        "actor": log.actor,
        "metadata": log.metadata_payload or {},
        "created_display": log.created_at.strftime("%d.%m.%Y %H:%M"),
    }


def serialize_catalog_entry(entry: ProductCatalogEntry) -> dict[str, Any]:
    return {
        "id": entry.id,
        "sku": entry.sku or "",
        "department": entry.department or "",
        "usage_area": entry.usage_area.name if entry.usage_area else "",
        "license_name": entry.license_name.name if entry.license_name else "",
        "info_category": entry.info_category.name if entry.info_category else "",
        "factory": entry.factory.name if entry.factory else "",
        "hardware_type": entry.hardware_type.name if entry.hardware_type else "",
        "brand": entry.brand.name if entry.brand else "",
        "model": entry.model.name if entry.model else "",
        "created_display": entry.created_at.strftime("%d.%m.%Y %H:%M"),
    }


def load_request_groups() -> dict[str, Any]:
    request_groups_payload: list[dict[str, Any]] = []
    groups = (
        RequestGroup.query.options(
            joinedload(RequestGroup.orders)
            .joinedload(RequestOrder.lines)
            .joinedload(RequestLine.order),
            joinedload(RequestGroup.orders).joinedload(RequestOrder.snapshots),
        )
        .order_by(RequestGroup.id)
        .all()
    )

    for group in groups:
        orders_payload = [serialize_request_order(order) for order in group.orders]
        request_groups_payload.append(
            {
                "key": group.key,
                "label": group.label,
                "description": group.description,
                "empty_message": group.empty_message,
                "orders": orders_payload,
            }
        )

    brands = Brand.query.options(joinedload(Brand.models)).order_by(Brand.name).all()

    models_by_brand: dict[str, list[str]] = {}
    for brand in brands:
        models_by_brand[brand.name] = [model.name for model in brand.models]

    hardware_catalog = {
        "types": [ht.name for ht in HardwareType.query.order_by(HardwareType.name)],
        "brands": [brand.name for brand in brands],
        "models": [
            model.name for model in HardwareModel.query.order_by(HardwareModel.name)
        ],
        "models_by_brand": models_by_brand,
    }

    return {
        "request_groups": request_groups_payload,
        "hardware_catalog": hardware_catalog,
        "stock_metadata_config": STOCK_METADATA_FIELDS,
        "stock_support_options": build_stock_support_options(),
        "stock_category_labels": STOCK_CATEGORY_LABELS,
    }


def get_inventory_item_with_relations(item_id: int) -> InventoryItem | None:
    return InventoryItem.query.options(
        joinedload(InventoryItem.factory),
        joinedload(InventoryItem.hardware_type),
        joinedload(InventoryItem.brand),
        joinedload(InventoryItem.model),
        joinedload(InventoryItem.responsible_user),
        joinedload(InventoryItem.events),
        joinedload(InventoryItem.licenses),
        joinedload(InventoryItem.maintenances),
    ).get(item_id)


def get_stock_item_with_relations(item_id: int) -> StockItem | None:
    return (
        StockItem.query.options(
            joinedload(StockItem.inventory_item).joinedload(
                InventoryItem.hardware_type
            ),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.factory),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.brand),
            joinedload(StockItem.inventory_item).joinedload(InventoryItem.model),
            joinedload(StockItem.license),
            joinedload(StockItem.category_ref),
            joinedload(StockItem.unit_ref),
            joinedload(StockItem.logs),
        )
        .filter_by(id=item_id, is_deleted=False)
        .first()
    )


def get_request_order_with_relations(order_id: int) -> RequestOrder | None:
    return (
        RequestOrder.query.options(
            joinedload(RequestOrder.group),
            joinedload(RequestOrder.lines),
            joinedload(RequestOrder.snapshots),
        )
        .filter_by(id=order_id)
        .first()
    )


def get_request_group_by_key(key: str) -> RequestGroup | None:
    normalized = (key or "").strip().lower()
    if not normalized:
        return None
    return RequestGroup.query.filter(func.lower(RequestGroup.key) == normalized).first()


def add_inventory_event(
    item: InventoryItem,
    event_type: str,
    note: str | None = None,
    performed_by: str = DEFAULT_EVENT_ACTOR,
) -> InventoryEvent:
    event = InventoryEvent(
        item=item,
        event_type=event_type,
        performed_by=performed_by,
        note=note or None,
    )
    db.session.add(event)
    record_activity(
        area="envanter",
        action=event_type,
        description=note,
        actor=performed_by,
        metadata={
            "inventory_id": item.id,
            "inventory_no": item.inventory_no,
            "status": item.status,
        },
    )
    return event


def resolve_stock_category(name: str | None) -> StockCategory | None:
    cleaned = (name or "").strip()
    if not cleaned:
        return None
    existing = find_existing_by_name(StockCategory, cleaned)
    if existing:
        return existing
    category = StockCategory(name=cleaned)
    db.session.add(category)
    db.session.flush()
    return category


def resolve_stock_unit(name: str | None) -> StockUnit | None:
    cleaned = (name or "").strip()
    if not cleaned:
        return None
    existing = find_existing_by_name(StockUnit, cleaned)
    if existing:
        return existing
    unit = StockUnit(name=cleaned)
    db.session.add(unit)
    db.session.flush()
    return unit


def record_stock_movement(
    stock_item: StockItem,
    *,
    operation_type: str,
    old_quantity: int,
    new_quantity: int,
    user: User | None,
) -> StockMovement:
    movement = StockMovement(
        stock_item=stock_item,
        user_id=user.id if user else None,
        operation_type=operation_type,
        old_quantity=max(0, int(old_quantity)),
        new_quantity=max(0, int(new_quantity)),
    )
    db.session.add(movement)
    return movement


def record_stock_audit(
    stock_item: StockItem,
    *,
    old_quantity: int,
    new_quantity: int,
    performed_by: str,
) -> StockAuditLog:
    audit = StockAuditLog(
        stock_item=stock_item,
        old_quantity=max(0, int(old_quantity)),
        new_quantity=max(0, int(new_quantity)),
        performed_by=(performed_by or DEFAULT_EVENT_ACTOR).strip()
        or DEFAULT_EVENT_ACTOR,
    )
    db.session.add(audit)
    return audit


def record_stock_log(
    stock_item: StockItem,
    action: str,
    *,
    action_type: str = "info",
    performed_by: str | None = None,
    quantity_change: int = 0,
    note: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> StockLog:
    actor = (performed_by or DEFAULT_EVENT_ACTOR).strip() or DEFAULT_EVENT_ACTOR
    log = StockLog(
        stock_item=stock_item,
        action=action,
        action_type=action_type,
        performed_by=actor,
        quantity_change=quantity_change,
        note=note or None,
    )
    log.metadata_payload = metadata or None
    db.session.add(log)

    activity_metadata = {
        "stock_item_id": stock_item.id,
        "stock_item_title": stock_item.title,
        "stock_item_status": stock_item.status,
    }
    if metadata:
        activity_metadata.update(metadata)

    record_activity(
        area="stok",
        action=action,
        description=note or stock_item.title,
        actor=actor,
        metadata=activity_metadata,
    )
    return log


def build_inventory_stock_metadata(item: InventoryItem) -> dict[str, str]:
    return {
        "inventory_no": item.inventory_no or "",
        "computer_name": item.computer_name or "",
        "hostname": item.computer_name or "",
        "factory": item.factory.name if item.factory else "",
        "department": item.department or "",
        "hardware_type": item.hardware_type.name if item.hardware_type else "",
        "brand": item.brand.name if item.brand else "",
        "model": item.model.name if item.model else "",
        "serial_no": item.serial_no or "",
        "ifs_no": item.ifs_no or "",
        "ip_address": item.related_machine_no or "",
        "mac_address": item.machine_no or "",
        "responsible": (
            f"{item.responsible_user.first_name} {item.responsible_user.last_name}"
            if item.responsible_user
            else ""
        ),
    }


def create_stock_item_from_inventory(
    item: InventoryItem,
    *,
    note: str | None = None,
    actor: str = DEFAULT_EVENT_ACTOR,
) -> StockItem:
    title_parts = [
        item.brand.name if item.brand else "",
        item.model.name if item.model else "",
    ]
    title = " ".join(part for part in title_parts if part).strip()
    if not title:
        title = item.inventory_no or "Envanter"

    category_value = determine_stock_category_from_inventory(item)

    stock_item = StockItem(
        source_type="inventory",
        inventory_item=item,
        reference_code=item.inventory_no,
        title=title,
        category=category_value,
        quantity=1,
        status="stokta",
        note=note or None,
    )
    metadata_payload = build_inventory_stock_metadata(item)
    metadata_payload = remove_assignment_only_metadata(metadata_payload, category_value)
    stock_item.metadata_payload = {
        key: value for key, value in metadata_payload.items() if value
    }
    db.session.add(stock_item)
    db.session.flush()
    record_stock_log(
        stock_item,
        "Stok girişi",
        action_type="in",
        performed_by=actor,
        quantity_change=1,
        note=note,
        metadata={"inventory_no": item.inventory_no},
    )
    record_stock_movement(
        stock_item,
        operation_type="giris",
        old_quantity=0,
        new_quantity=1,
        user=get_active_user(),
    )
    return stock_item


def create_stock_item_from_license(
    license: InventoryLicense,
    *,
    note: str | None = None,
    actor: str = DEFAULT_EVENT_ACTOR,
) -> StockItem:
    display_name, key = split_license_name(license.name)
    title = display_name or license.name
    stock_item = StockItem(
        source_type="license",
        license=license,
        reference_code=license.name,
        title=title or "Lisans",
        category="lisans",
        quantity=1,
        status="stokta",
        note=note or None,
    )
    associated_item = license.item
    stock_item.metadata_payload = {
        "license_key": key,
        "license_name": title,
        "inventory_no": associated_item.inventory_no if associated_item else "",
        "department": associated_item.department if associated_item else "",
        "factory": (
            associated_item.factory.name
            if associated_item and associated_item.factory
            else ""
        ),
    }
    db.session.add(stock_item)
    db.session.flush()
    record_stock_log(
        stock_item,
        "Lisans stok girişi",
        action_type="in",
        performed_by=actor,
        quantity_change=1,
        note=note,
        metadata={"license_id": license.id},
    )
    record_stock_movement(
        stock_item,
        operation_type="giris",
        old_quantity=0,
        new_quantity=1,
        user=get_active_user(),
    )
    return stock_item


def create_stock_item_from_request_line(
    order: RequestOrder,
    line: RequestLine,
    *,
    quantity: int,
    note: str | None = None,
    actor: str = DEFAULT_EVENT_ACTOR,
    category: str | None = None,
    metadata: dict[str, str] | None = None,
) -> StockItem:
    title_parts = [line.brand, line.model]
    title = " ".join(part for part in title_parts if part).strip() or line.hardware_type
    category_value = normalize_stock_category(
        category or line.category,
        fallback="talep",
    )
    metadata_payload = {
        "request_no": order.order_no,
        "department": order.department,
        "hardware_type": line.hardware_type,
        "brand": line.brand,
        "model": line.model,
    }
    if metadata:
        metadata_payload.update(metadata)
    reference_code = (
        metadata_payload.get("inventory_no")
        or metadata_payload.get("license_key")
        or order.order_no
    )
    stock_item = StockItem(
        source_type="request",
        source_id=order.id,
        reference_code=reference_code,
        title=title or "Talep Öğesi",
        category=category_value,
        quantity=max(1, quantity),
        status="stokta",
        note=note or None,
    )
    stock_item.metadata_payload = metadata_payload
    db.session.add(stock_item)
    db.session.flush()
    log_metadata = dict(metadata_payload)
    log_metadata["request_id"] = order.id
    record_stock_log(
        stock_item,
        "Talep stok girişi",
        action_type="in",
        performed_by=actor,
        quantity_change=stock_item.quantity,
        note=note,
        metadata=log_metadata,
    )
    record_stock_movement(
        stock_item,
        operation_type="giris",
        old_quantity=0,
        new_quantity=stock_item.quantity,
        user=get_active_user(),
    )
    return stock_item


def prepare_stock_metadata(
    category: str,
    payload: Any,
    *,
    defaults: dict[str, Any] | None = None,
    include_assignment_fields: bool = True,
) -> dict[str, str]:
    schema = STOCK_METADATA_FIELDS.get(category, [])
    if not include_assignment_fields:
        schema = [field for field in schema if not field.get("assignment_only")]
    provided: dict[str, Any]
    if isinstance(payload, dict):
        provided = payload
    else:
        provided = {}
    defaults = defaults or {}
    cleaned: dict[str, str] = {}

    def normalize_value(raw: Any) -> str:
        if raw is None:
            return ""
        if isinstance(raw, str):
            return raw.strip()
        return str(raw).strip()

    for field in schema:
        key = field["key"]
        label = field.get("label", key.capitalize())
        value = normalize_value(provided.get(key))
        if not value:
            value = normalize_value(defaults.get(key))
        if not value and field.get("required"):
            raise ValueError(f"{label} alanı zorunludur.")
        if value:
            cleaned[key] = value

    for key, value in provided.items():
        if key in cleaned:
            continue
        normalized = normalize_value(value)
        if normalized:
            cleaned[key] = normalized

    return cleaned


def json_error(message: str) -> dict[str, str]:
    return {"error": message}


def record_activity(
    *,
    area: str,
    action: str,
    description: str | None = None,
    actor: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> ActivityLog:
    log = ActivityLog(
        area=area,
        action=action,
        description=description or None,
        actor=actor or DEFAULT_EVENT_ACTOR,
        metadata_payload=metadata or None,
    )
    db.session.add(log)
    return log


def load_activity_logs(limit: int | None = None) -> list[dict[str, Any]]:
    query = ActivityLog.query.order_by(ActivityLog.created_at.desc())
    if limit is not None:
        query = query.limit(limit)
    return [serialize_activity_log(log) for log in query.all()]


def load_recent_activity(limit: int = 6) -> list[dict[str, Any]]:
    allowed_areas = {
        "talep",
        "urun",
        "kullanici",
        "envanter",
        "stok",
        "bilgi",
        "profil",
        "auth",
        "sistem",
        "entegrasyon",
    }
    query_limit = max(limit * 4, limit)
    candidates = (
        ActivityLog.query.order_by(ActivityLog.created_at.desc())
        .limit(query_limit)
        .all()
    )
    filtered: list[dict[str, Any]] = []
    for log in candidates:
        if log.area not in allowed_areas:
            continue
        filtered.append(serialize_activity_log(log))
        if len(filtered) >= limit:
            break
    return filtered


def maintenance_candidate_items_query():
    return InventoryItem.query.options(
        joinedload(InventoryItem.hardware_type),
        joinedload(InventoryItem.maintenances),
    )


def calculate_maintenance_status(
    last_maintenance_at: date | datetime | None, today: date | None = None
) -> dict[str, Any]:
    check_date = today or datetime.utcnow().date()
    if last_maintenance_at is None:
        return {
            "last_maintenance_display": "-",
            "days_since_maintenance": None,
            "days_until_due": None,
            "status": "none",
            "label": "Bakım kaydı yok",
        }

    maintenance_date = (
        last_maintenance_at.date()
        if isinstance(last_maintenance_at, datetime)
        else last_maintenance_at
    )
    days_since_maintenance = (check_date - maintenance_date).days
    days_until_due = 365 - days_since_maintenance

    if days_since_maintenance >= 365:
        status = "overdue"
        label = "Bakım gecikti"
    elif 335 <= days_since_maintenance < 365:
        status = "warning"
        label = "1 ay içinde bakım"
    else:
        status = "ok"
        label = "Güncel"

    return {
        "last_maintenance_display": maintenance_date.strftime("%d.%m.%Y"),
        "days_since_maintenance": days_since_maintenance,
        "days_until_due": days_until_due,
        "status": status,
        "label": label,
    }


def maintenance_status_badge_class(status: str) -> str:
    if status in {"overdue", "none"}:
        return "maintenance-badge-overdue"
    if status == "warning":
        return "maintenance-badge-warning"
    return "text-bg-success-subtle text-success"


def maintenance_row_class(status: str) -> str:
    if status in {"overdue", "none"}:
        return "maintenance-row-overdue"
    if status == "warning":
        return "maintenance-row-warning"
    return ""


def load_maintenance_dashboard_counts() -> dict[str, int]:
    due_count = 0
    warning_count = 0
    items = maintenance_candidate_items_query().all()
    for item in items:
        if not is_computer_hardware_type(
            item.hardware_type.name if item.hardware_type else None
        ):
            continue
        if (item.status or "").lower() in {"hurda", "stokta"}:
            continue

        last_maintenance = item.maintenances[0] if item.maintenances else None
        maintenance_status = calculate_maintenance_status(
            last_maintenance.performed_at if last_maintenance else None
        )["status"]
        if maintenance_status in {"overdue", "none"}:
            due_count += 1
        elif maintenance_status == "warning":
            warning_count += 1

    return {
        "maintenance_due_count": due_count,
        "maintenance_warning_count": warning_count,
    }


def load_dashboard_metrics() -> dict[str, Any]:
    available_stock = (
        db.session.query(func.sum(StockItem.quantity))
        .filter(StockItem.status == "stokta")
        .scalar()
        or 0
    )
    total_stock = db.session.query(func.sum(StockItem.quantity)).scalar() or 0

    open_request_count = (
        RequestOrder.query.join(RequestGroup)
        .filter(func.lower(RequestGroup.key) == "acik")
        .count()
    )
    total_request_count = RequestOrder.query.count()

    faulty_inventory_count = InventoryItem.query.filter(
        InventoryItem.status == "arizali"
    ).count()
    critical_stock_count = StockItem.query.filter(
        StockItem.status.in_(["arizali", "hurda"])
    ).count()
    maintenance_counts = load_maintenance_dashboard_counts()
    recent_stock_movements = (
        StockMovement.query.options(joinedload(StockMovement.stock_item))
        .order_by(StockMovement.created_at.desc())
        .limit(5)
        .all()
    )

    critical_alerts = (
        faulty_inventory_count
        + critical_stock_count
        + maintenance_counts["maintenance_due_count"]
    )

    return {
        "available_stock": int(available_stock),
        "total_stock": int(total_stock),
        "open_requests": int(open_request_count),
        "total_requests": int(total_request_count),
        "critical_alerts": int(critical_alerts),
        "faulty_inventory": int(faulty_inventory_count),
        "problem_stock": int(critical_stock_count),
        "maintenance_due_count": int(maintenance_counts["maintenance_due_count"]),
        "maintenance_warning_count": int(
            maintenance_counts["maintenance_warning_count"]
        ),
        "recent_stock_movements": [
            {
                "operation": movement.operation_type,
                "title": movement.stock_item.title if movement.stock_item else "Stok",
                "created_display": (
                    movement.created_at.strftime("%d.%m.%Y %H:%M")
                    if movement.created_at
                    else ""
                ),
            }
            for movement in recent_stock_movements
        ],
    }


def build_stock_support_options() -> dict[str, list[str]]:
    factory_names = [factory.name for factory in Factory.query.order_by(Factory.name)]

    department_values: set[str] = set()
    for (department,) in db.session.query(InventoryItem.department).distinct():
        if department:
            department_values.add(department)
    for (department,) in db.session.query(User.department).distinct():
        if department:
            department_values.add(department)
    department_names = sorted(department_values)

    responsible_names = [
        f"{user.first_name} {user.last_name}".strip()
        for user in active_users_query().order_by(User.first_name, User.last_name)
        if (user.first_name or user.last_name)
    ]

    usage_area_names = [
        usage_area.name for usage_area in UsageArea.query.order_by(UsageArea.name)
    ]
    license_name_values = [
        license_name.name
        for license_name in LicenseName.query.order_by(LicenseName.name)
    ]

    inventory_numbers = [
        inventory_no
        for (inventory_no,) in db.session.query(InventoryItem.inventory_no)
        .filter(InventoryItem.inventory_no.isnot(None))
        .distinct()
        .order_by(InventoryItem.inventory_no)
    ]
    return {
        "factories": factory_names,
        "departments": department_names,
        "responsibles": responsible_names,
        "usage_areas": usage_area_names,
        "license_names": license_name_values,
        "inventory_numbers": inventory_numbers,
    }


def load_admin_panel_payload() -> dict:
    users = (
        active_users_query(include_inactive=True)
        .order_by(User.first_name, User.last_name)
        .all()
    )
    stock_support_options = build_stock_support_options()
    department_options = [
        {"id": name, "name": name}
        for name in stock_support_options.get("departments", [])
    ]

    product_options = {
        "usage_areas": [
            ua.to_dict() for ua in UsageArea.query.order_by(UsageArea.name)
        ],
        "license_names": [
            ln.to_dict() for ln in LicenseName.query.order_by(LicenseName.name)
        ],
        "info_categories": [
            ic.to_dict() for ic in InfoCategory.query.order_by(InfoCategory.name)
        ],
        "factories": [
            factory.to_dict() for factory in Factory.query.order_by(Factory.name)
        ],
        "hardware_types": [
            ht.to_dict() for ht in HardwareType.query.order_by(HardwareType.name)
        ],
        "brands": [brand.to_dict() for brand in Brand.query.order_by(Brand.name)],
        "departments": department_options,
    }

    brand_models = [
        brand.to_dict(include_models=True) for brand in Brand.query.order_by(Brand.name)
    ]
    ldap_profiles = [
        profile.to_dict() for profile in LdapProfile.query.order_by(LdapProfile.name)
    ]
    catalog_entries = (
        ProductCatalogEntry.query.options(
            joinedload(ProductCatalogEntry.usage_area),
            joinedload(ProductCatalogEntry.license_name),
            joinedload(ProductCatalogEntry.info_category),
            joinedload(ProductCatalogEntry.factory),
            joinedload(ProductCatalogEntry.hardware_type),
            joinedload(ProductCatalogEntry.brand),
            joinedload(ProductCatalogEntry.model),
        )
        .filter(ProductCatalogEntry.is_deleted.is_(False))
        .order_by(ProductCatalogEntry.created_at.desc())
        .all()
    )

    return {
        "users": users,
        "product_options": product_options,
        "brand_models": brand_models,
        "ldap_profiles": ldap_profiles,
        "catalog_entries": [
            serialize_catalog_entry(entry) for entry in catalog_entries
        ],
    }


def parse_option_name(payload: dict | None) -> str:
    if not payload or not isinstance(payload, dict):
        raise ValueError("Geçersiz istek gövdesi")
    name = (payload.get("name") or "").strip()
    if not name:
        raise ValueError("İsim alanı zorunludur")
    return name


def parse_ldap_profile_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("Geçersiz istek gövdesi")

    def clean_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    name = clean_text(payload.get("name"))
    host = clean_text(payload.get("host"))
    base_dn = clean_text(payload.get("base_dn"))
    bind_dn = clean_text(payload.get("bind_dn"))
    port_value = payload.get("port")
    try:
        port = int(port_value)
    except (TypeError, ValueError):
        port = 0

    if not name:
        raise ValueError("Profil adı zorunludur.")
    if not host:
        raise ValueError("Sunucu adresi zorunludur.")
    if port <= 0:
        raise ValueError("Geçerli bir port numarası girin.")
    if not base_dn:
        raise ValueError("Base DN alanı zorunludur.")
    if not bind_dn:
        raise ValueError("Bind kullanıcı alanı zorunludur.")

    return {
        "name": name,
        "host": host,
        "port": port,
        "base_dn": base_dn,
        "bind_dn": bind_dn,
    }


def create_brand():
    try:
        name = parse_option_name(request.get_json())
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    if find_existing_by_name(Brand, name):
        return jsonify({"error": "Bu marka zaten mevcut."}), 409

    brand = Brand(name=name)
    db.session.add(brand)
    db.session.commit()
    return jsonify(brand.to_dict(include_models=True)), 201


def delete_brand(brand_id: int):
    brand = Brand.query.get(brand_id)
    if brand is None:
        return jsonify({"error": "Marka bulunamadı."}), 404

    db.session.delete(brand)
    db.session.commit()
    return ("", 204)


OPTION_MODEL_MAPPING = {
    "usage-areas": UsageArea,
    "license-names": LicenseName,
    "info-categories": InfoCategory,
    "factories": Factory,
    "hardware-types": HardwareType,
}


def seed_initial_data() -> None:
    seed_simple_users()
    seed_product_metadata()
    seed_information_entries()
    seed_inventory_data()
    seed_ldap_profiles()
    seed_request_data()
    seed_stock_reference_data()
    seed_stock_data()
    db.session.commit()


def seed_simple_users() -> None:
    existing_user_count = User.query.count()

    admin_password = generate_password_hash("admin")
    created_users: list[User] = []

    admin_user = User.query.filter(func.lower(User.username) == "admin").first()

    if admin_user is None:
        admin_user = User(
            username="admin",
            first_name="Stok",
            last_name="Yöneticisi",
            email="admin@example.com",
            role="Sistem Süper Yöneticisi",
            department="Bilgi Teknolojileri",
            password_hash=admin_password,
            system_role="superadmin",
            must_change_password=True,
        )
        db.session.add(admin_user)
        created_users.append(admin_user)
    else:
        updated = False
        if not admin_user.password_hash:
            admin_user.password_hash = admin_password
            updated = True
        if not admin_user.system_role or admin_user.system_role.lower() not in {
            "admin",
            "superadmin",
        }:
            admin_user.system_role = "superadmin"
            updated = True
        if admin_user.must_change_password is None:
            admin_user.must_change_password = True
            updated = True
        if updated:
            created_users.append(admin_user)

    if existing_user_count:
        if created_users:
            record_activity(
                area="kullanici",
                action="Varsayılan yönetici güncellendi",
                description="Eksik yönetici hesabı oluşturuldu veya güncellendi.",
                metadata={"count": len(created_users)},
            )
        return

    default_password = generate_password_hash("Parola123!")
    demo_users = [
        User(
            username="m.cetin",
            first_name="Merve",
            last_name="Çetin",
            email="merve.cetin@example.com",
            role="Yönetici",
            department="IT Operasyon",
            password_hash=default_password,
            system_role="admin",
        ),
        User(
            username="a.kaya",
            first_name="Ahmet",
            last_name="Kaya",
            email="ahmet.kaya@example.com",
            role="Satın Alma Uzmanı",
            department="Satın Alma",
            password_hash=default_password,
            system_role="user",
        ),
        User(
            username="z.ucar",
            first_name="Zeynep",
            last_name="Uçar",
            email="zeynep.ucar@example.com",
            role="Depo Sorumlusu",
            department="Lojistik",
            password_hash=default_password,
            system_role="user",
        ),
        User(
            username="b.tan",
            first_name="Berk",
            last_name="Tan",
            email="berk.tan@example.com",
            role="Destek Uzmanı",
            department="Teknik Destek",
            password_hash=default_password,
            system_role="user",
        ),
        User(
            username="e.sonmez",
            first_name="Elif",
            last_name="Sönmez",
            email="elif.sonmez@example.com",
            role="Finans Analisti",
            department="Finans",
            password_hash=default_password,
            system_role="user",
        ),
    ]

    db.session.add_all(demo_users)
    created_users.extend(demo_users)

    record_activity(
        area="kullanici",
        action="Varsayılan kullanıcılar eklendi",
        description="Sistem başlangıç kullanıcıları oluşturuldu.",
        metadata={"count": len(created_users)},
    )


def seed_product_metadata() -> None:
    added_any = False

    if not UsageArea.query.count():
        db.session.add_all(
            UsageArea(name=name)
            for name in ["Ofis", "Saha", "Veri Merkezi", "Üretim", "Uzaktan Çalışma"]
        )
        added_any = True

    if not LicenseName.query.count():
        db.session.add_all(
            LicenseName(name=name)
            for name in [
                "Microsoft 365 Business",
                "Adobe Creative Cloud",
                "JetBrains All Products",
                "AutoCAD LT",
            ]
        )
        added_any = True

    if not InfoCategory.query.count():
        db.session.add_all(
            InfoCategory(name=name)
            for name in ["Güvenlik", "İş Uygulamaları", "İletişim", "Altyapı"]
        )
        added_any = True

    if not Factory.query.count():
        db.session.add_all(
            Factory(name=name)
            for name in [
                "İstanbul Merkez",
                "Ankara Veri Merkezi",
                "İzmir Üretim",
                "Bursa Lojistik",
            ]
        )
        added_any = True

    if not HardwareType.query.count():
        db.session.add_all(
            HardwareType(name=name)
            for name in [
                "Laptop",
                "Masaüstü",
                "Monitör",
                "Sunucu",
                "Yazıcı",
                "Tarayıcı",
                "Tablet",
                "Aksesuar",
            ]
        )
        added_any = True

    if not Brand.query.count():
        brand_seed = {
            "Apple": ["MacBook Pro 14", "MacBook Air M2", "iMac 24"],
            "Asus": ["ZenBook 14", "ROG Zephyrus G14"],
            "Dell": ["Latitude 5440", "XPS 15", "PowerEdge R750"],
            "Fujitsu": ["fi-7160"],
            "HP": ["ProBook 450 G10", "EliteBook 840", "LaserJet Pro M404"],
            "Lenovo": ["ThinkPad X1 Carbon", "ThinkSystem SR250"],
            "Samsung": ["Galaxy Book3", "ViewFinity S8"],
        }
        for brand_name, models in brand_seed.items():
            brand = Brand(name=brand_name)
            brand.models = [HardwareModel(name=model_name) for model_name in models]
            db.session.add(brand)
        added_any = True

    if added_any:
        record_activity(
            area="urun",
            action="Ürün katalog seçenekleri hazırlandı",
            description="Varsayılan marka, model ve kullanım alanı verileri yüklendi.",
        )


def seed_information_entries() -> None:
    if InfoEntry.query.count():
        return

    categories = {category.name: category for category in InfoCategory.query.all()}

    sample_entries = [
        {
            "title": "Sosyal Mühendislik Farkındalığı",
            "category": "Güvenlik",
            "content": (
                "Şüpheli e-posta ve bağlantıları bildirmeden açmayın. Kurumsal sistemlere erişim "
                "sağlarken her zaman çok faktörlü kimlik doğrulamayı kullanın."
            ),
        },
        {
            "title": "VPN Kullanım Kılavuzu",
            "category": "Altyapı",
            "content": (
                "Uzak bağlantı kurmadan önce cihazınızın güncel olduğundan emin olun ve bağlantı "
                "esnasında sadece iş amaçlı kaynaklara erişin."
            ),
        },
        {
            "title": "Yeni Satın Alma Süreçleri",
            "category": "İş Uygulamaları",
            "content": (
                "Tüm donanım talepleri Talep Takip sayfası üzerinden açılmalı ve satın alma onayı "
                "alınmadan sipariş verilmemelidir."
            ),
        },
    ]

    created_count = 0
    for payload in sample_entries:
        category = categories.get(payload["category"])
        if not category:
            continue
        entry = InfoEntry(
            title=payload["title"],
            category=category,
            content=payload["content"],
        )
        db.session.add(entry)
        created_count += 1

    if created_count:
        record_activity(
            area="bilgi",
            action="Bilgi kayıtları oluşturuldu",
            description="Varsayılan bilgi içerikleri eklendi.",
            metadata={"count": created_count},
        )


def seed_inventory_data() -> None:
    if InventoryItem.query.count():
        return

    factories = {factory.name: factory for factory in Factory.query.all()}
    hardware_types = {ht.name: ht for ht in HardwareType.query.all()}
    users = {f"{user.first_name} {user.last_name}": user for user in User.query.all()}
    brands = {
        brand.name: brand
        for brand in Brand.query.options(joinedload(Brand.models)).all()
    }

    model_lookup = {}
    for brand in brands.values():
        for model in brand.models:
            model_lookup[(brand.name, model.name)] = model

    now = datetime.utcnow()

    item_primary = InventoryItem(
        inventory_no="ENV-000123",
        computer_name="PC-OFIS-01",
        factory=factories.get("İstanbul Merkez"),
        department="IT Operasyon",
        hardware_type=hardware_types.get("Laptop"),
        responsible_user=users.get("Ahmet Kaya"),
        brand=brands.get("Dell"),
        model=model_lookup.get(("Dell", "Latitude 5440")),
        serial_no="SN123456789",
        ifs_no="IFS-00045",
        related_machine_no="",
        machine_no="PC-LAP-01",
        note="IT destek ekibine teslim edildi.",
        status="aktif",
    )
    item_primary.licenses = [
        InventoryLicense(name="Office 2021 - 123456789", status="aktif"),
        InventoryLicense(name="Visio Professional - 987654321", status="aktif"),
    ]
    item_primary.events = [
        InventoryEvent(
            event_type="Stok Girişi",
            performed_by="Berk Tan",
            performed_at=now - timedelta(days=120),
            note="Merkez depoya giriş yapıldı.",
        ),
        InventoryEvent(
            event_type="Atama",
            performed_by="Merve Çetin",
            performed_at=now - timedelta(days=90),
            note="Cihaz Ahmet Kaya'ya teslim edildi.",
        ),
        InventoryEvent(
            event_type="Bakım",
            performed_by="Zeynep Uçar",
            performed_at=now - timedelta(days=15),
            note="Genel bakım ve temizlik yapıldı.",
        ),
    ]

    item_faulty = InventoryItem(
        inventory_no="ENV-000207",
        computer_name="PC-LOG-03",
        factory=factories.get("Bursa Lojistik"),
        department="Lojistik",
        hardware_type=hardware_types.get("Monitör"),
        responsible_user=users.get("Zeynep Uçar"),
        brand=brands.get("Samsung"),
        model=model_lookup.get(("Samsung", "ViewFinity S8")),
        serial_no="SN987654321",
        ifs_no="IFS-00112",
        related_machine_no="LOG-WS-04",
        machine_no="MN-LOG-03",
        note="Ekran arızası nedeniyle servise gönderilecek.",
        status="arizali",
    )
    item_faulty.licenses = [
        InventoryLicense(name="Adobe Creative Cloud - LZ-55981", status="aktif"),
    ]
    item_faulty.events = [
        InventoryEvent(
            event_type="Atama",
            performed_by="Merve Çetin",
            performed_at=now - timedelta(days=200),
            note="Zeynep Uçar'a teslim edildi.",
        ),
        InventoryEvent(
            event_type="Arıza Bildirimi",
            performed_by="Zeynep Uçar",
            performed_at=now - timedelta(days=7),
            note="Ekranda titreme sorunu bildirildi.",
        ),
        InventoryEvent(
            event_type="Tamir",
            performed_by="Servis Sağlayıcısı",
            performed_at=now - timedelta(days=2),
            note="Parça siparişi bekleniyor.",
        ),
    ]

    printer_central = InventoryItem(
        inventory_no="PRN-000444",
        computer_name="PRN-MERKEZ-01",
        factory=factories.get("İstanbul Merkez"),
        department="IT Operasyon",
        hardware_type=hardware_types.get("Yazıcı"),
        responsible_user=users.get("Merve Çetin"),
        brand=brands.get("HP"),
        model=model_lookup.get(("HP", "LaserJet Pro M404")),
        serial_no="HP444MERKEZ",
        ifs_no="IFS-00444",
        related_machine_no="10.0.0.32",
        machine_no="AA:BC:44:32:10:01",
        note="Merkez ofiste paylaşımlı yazıcı olarak kullanılıyor.",
        status="aktif",
    )
    printer_central.events = [
        InventoryEvent(
            event_type="Stok Girişi",
            performed_by="Berk Tan",
            performed_at=now - timedelta(days=60),
            note="Merkez depoya teslim alındı.",
        ),
        InventoryEvent(
            event_type="Atama",
            performed_by="Merve Çetin",
            performed_at=now - timedelta(days=58),
            note="IT Operasyon ekibine paylaşımlı olarak tanımlandı.",
        ),
        InventoryEvent(
            event_type="Bakım",
            performed_by="Servis Sağlayıcısı",
            performed_at=now - timedelta(days=12),
            note="Toner ve drum değişimi yapıldı.",
        ),
    ]

    printer_faulty = InventoryItem(
        inventory_no="PRN-000558",
        computer_name="PRN-LOG-01",
        factory=factories.get("Bursa Lojistik"),
        department="Lojistik",
        hardware_type=hardware_types.get("Yazıcı"),
        responsible_user=users.get("Zeynep Uçar"),
        brand=brands.get("HP"),
        model=model_lookup.get(("HP", "LaserJet Pro M404")),
        serial_no="HP558LOGISTIK",
        ifs_no="IFS-00558",
        related_machine_no="10.0.0.78",
        machine_no="AA:BC:55:58:10:01",
        note="Kağıt besleme ünitesinde sıkışma sorunu gözlemlendi.",
        status="arizali",
    )
    printer_faulty.events = [
        InventoryEvent(
            event_type="Atama",
            performed_by="Ahmet Kaya",
            performed_at=now - timedelta(days=180),
            note="Lojistik depoya kurulum yapıldı.",
        ),
        InventoryEvent(
            event_type="Arıza Bildirimi",
            performed_by="Zeynep Uçar",
            performed_at=now - timedelta(days=3),
            note="Kağıt besleme ünitesi kontrol edilmek üzere servis çağırıldı.",
        ),
    ]

    item_retired = InventoryItem(
        inventory_no="ENV-000318",
        computer_name="PRN-FN-02",
        factory=factories.get("Ankara Veri Merkezi"),
        department="Finans",
        hardware_type=hardware_types.get("Yazıcı"),
        responsible_user=users.get("Elif Sönmez"),
        brand=brands.get("HP"),
        model=model_lookup.get(("HP", "LaserJet Pro M404")),
        serial_no="SN564738291",
        ifs_no="IFS-00221",
        related_machine_no="10.0.0.45",
        machine_no="AA:BC:31:18:00:02",
        note="Yeni yazıcı alındığından hurdaya ayrıldı.",
        status="hurda",
    )
    item_retired.licenses = [
        InventoryLicense(name="HP ePrint Service", status="pasif"),
    ]
    item_retired.events = [
        InventoryEvent(
            event_type="Stok Girişi",
            performed_by="Ahmet Kaya",
            performed_at=now - timedelta(days=400),
            note="Depoya giriş yapıldı.",
        ),
        InventoryEvent(
            event_type="Hurdaya Ayırma",
            performed_by="Elif Sönmez",
            performed_at=now - timedelta(days=5),
            note="Yeni model yazıcı ile değiştirildi.",
        ),
    ]

    db.session.add_all(
        [item_primary, item_faulty, printer_central, printer_faulty, item_retired]
    )
    record_activity(
        area="envanter",
        action="Örnek envanter kayıtları yüklendi",
        description="Sistem başlangıcı için örnek envanter kayıtları oluşturuldu.",
        metadata={"count": 5},
    )


def seed_ldap_profiles() -> None:
    if LdapProfile.query.count():
        return

    db.session.add_all(
        [
            LdapProfile(
                name="Merkez AD",
                host="ad.merkez.local",
                port=389,
                base_dn="DC=merkez,DC=local",
                bind_dn="CN=ldap.service,OU=Hizmet Hesaplari,DC=merkez,DC=local",
            ),
            LdapProfile(
                name="Uzak Ofis",
                host="ldap.uzakofis.local",
                port=636,
                base_dn="DC=uzakofis,DC=local",
                bind_dn="CN=ldap.reader,OU=Servis,DC=uzakofis,DC=local",
            ),
        ]
    )


def seed_request_data() -> None:
    if RequestGroup.query.count():
        return

    now = datetime.now()

    def make_order(
        *,
        group: RequestGroup,
        order_no: str,
        requested_by: str,
        department: str,
        opened_delta: timedelta,
        lines: list[dict],
    ) -> None:
        order = RequestOrder(
            order_no=order_no,
            requested_by=requested_by,
            department=department,
            opened_at=now - opened_delta,
            group=group,
        )
        for line in lines:
            order.lines.append(
                RequestLine(
                    hardware_type=line["hardware_type"],
                    brand=line["brand"],
                    model=line["model"],
                    quantity=line["quantity"],
                    note=line.get("note"),
                )
            )
        db.session.add(order)

    open_group = RequestGroup(
        key="acik",
        label="Açık",
        description="Açıkta bekleyen talepler buradan yönetilir.",
        empty_message="Bu statüde görüntülenecek açık talep bulunmuyor.",
    )
    db.session.add(open_group)
    make_order(
        group=open_group,
        order_no="SIP-2024-015",
        requested_by="Merve Çetin",
        department="IT Operasyon",
        opened_delta=timedelta(hours=2, minutes=45),
        lines=[
            {
                "hardware_type": "Laptop",
                "brand": "Dell",
                "model": "Latitude 5440",
                "quantity": 2,
                "note": "Saha ekibi için yedek cihazlar",
            },
            {
                "hardware_type": "Monitör",
                "brand": "Dell",
                "model": "P2422H",
                "quantity": 2,
                "note": "Yeni laptoplarla birlikte gönderilecek",
            },
        ],
    )
    make_order(
        group=open_group,
        order_no="SIP-2024-018",
        requested_by="Ahmet Kaya",
        department="Satın Alma",
        opened_delta=timedelta(days=1, hours=3),
        lines=[
            {
                "hardware_type": "Yazıcı",
                "brand": "HP",
                "model": "LaserJet Pro M404",
                "quantity": 1,
                "note": "Merkez ofis için yedek yazıcı",
            }
        ],
    )

    closed_group = RequestGroup(
        key="kapandi",
        label="Kapandı",
        description="Stoklara giren ve tamamlanan taleplerin özeti.",
        empty_message="Kapanmış talep kaydı bulunmuyor.",
    )
    db.session.add(closed_group)
    make_order(
        group=closed_group,
        order_no="SIP-2024-009",
        requested_by="Zeynep Uçar",
        department="Operasyon",
        opened_delta=timedelta(days=3, hours=5),
        lines=[
            {
                "hardware_type": "Sunucu",
                "brand": "Lenovo",
                "model": "ThinkSystem SR250",
                "quantity": 1,
                "note": "Veri merkezi genişletme talebi",
            }
        ],
    )
    make_order(
        group=closed_group,
        order_no="SIP-2024-011",
        requested_by="Berk Tan",
        department="Depo",
        opened_delta=timedelta(days=2, hours=8),
        lines=[
            {
                "hardware_type": "Tarayıcı",
                "brand": "Fujitsu",
                "model": "fi-7160",
                "quantity": 3,
                "note": "Yeni şube teslim alındı",
            }
        ],
    )

    cancelled_group = RequestGroup(
        key="iptal",
        label="İptal",
        description="İptal edilen talepler ve nedenlerine buradan ulaşabilirsiniz.",
        empty_message="İptal edilmiş talep kaydı bulunmuyor.",
    )
    db.session.add(cancelled_group)
    make_order(
        group=cancelled_group,
        order_no="SIP-2024-006",
        requested_by="Elif Sönmez",
        department="Finans",
        opened_delta=timedelta(days=5, hours=4),
        lines=[
            {
                "hardware_type": "Masaüstü",
                "brand": "HP",
                "model": "ProDesk 400",
                "quantity": 1,
                "note": "Bütçe onayı alınamadı",
            }
        ],
    )
    make_order(
        group=cancelled_group,
        order_no="SIP-2024-010",
        requested_by="Pelin Arı",
        department="Pazarlama",
        opened_delta=timedelta(days=4, hours=10),
        lines=[
            {
                "hardware_type": "Tablet",
                "brand": "Apple",
                "model": "iPad Air",
                "quantity": 4,
                "note": "Etkinlik ertelendiği için iptal edildi",
            }
        ],
    )

    total_orders = sum(
        len(group.orders) for group in (open_group, closed_group, cancelled_group)
    )
    record_activity(
        area="talep",
        action="Örnek talepler oluşturuldu",
        description="Açık, kapalı ve iptal statülerine örnek talepler eklendi.",
        metadata={"group_count": 3, "order_count": total_orders},
    )


def seed_stock_reference_data() -> None:
    for category_name in STOCK_CATEGORY_LABELS.keys():
        if not find_existing_by_name(StockCategory, category_name):
            db.session.add(StockCategory(name=category_name))
    for unit_name in ("adet", "kg", "metre"):
        if not find_existing_by_name(StockUnit, unit_name):
            db.session.add(StockUnit(name=unit_name))
    db.session.flush()


def seed_stock_data() -> None:
    if StockItem.query.count():
        return

    samples = [
        {
            "title": "Yedek Laptop Adaptörü",
            "category": "envanter",
            "quantity": 8,
            "note": "Saha ekipleri için hazır tutulan adaptörler.",
            "metadata": {"department": "IT Operasyon", "factory": "İstanbul Merkez"},
        },
        {
            "title": "HP 83A Toner",
            "category": "yazici",
            "quantity": 15,
            "note": "Merkez yazıcıları için stok toner.",
            "metadata": {"department": "Lojistik", "factory": "Bursa Lojistik"},
        },
        {
            "title": "Office 2021 Pro Plus",
            "category": "lisans",
            "quantity": 4,
            "note": "Yeni cihaz kurulumu için bekleyen lisans anahtarları.",
            "metadata": {"department": "IT Operasyon"},
        },
    ]

    for sample in samples:
        category_ref = resolve_stock_category(sample["category"])
        unit_ref = resolve_stock_unit("adet")
        stock_item = StockItem(
            source_type="manual",
            title=sample["title"],
            category=sample["category"],
            category_id=category_ref.id if category_ref else None,
            quantity=sample["quantity"],
            unit="adet",
            unit_id=unit_ref.id if unit_ref else None,
            status="stokta",
            note=sample["note"],
        )
        stock_item.metadata_payload = sample.get("metadata")
        db.session.add(stock_item)
        db.session.flush()
        record_stock_log(
            stock_item,
            "Başlangıç stok kaydı",
            action_type="in",
            performed_by="Sistem",
            quantity_change=stock_item.quantity,
            note=sample["note"],
        )


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)
